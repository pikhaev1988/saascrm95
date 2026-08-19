"""Импорт справочника заданий ВПР из JSON / Excel / CSV."""

from __future__ import annotations

import csv
import io
import json
import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from apps.vpr.exceptions import VprCatalogImportError
from apps.vpr.models import VprTaskCatalogEntry, VprTaskCatalogImport, VprTaskCatalogImportStatus
from apps.vpr.services.catalog_lookup import parse_task_code

logger = logging.getLogger(__name__)

FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "academic_year": ("academic_year", "year", "учебный_год", "учебный год", "год"),
    "subject": ("subject", "предмет"),
    "parallel": ("parallel", "class", "класс", "grade"),
    "task_number": ("task_number", "number", "номер", "номер_задания", "номер задания"),
    "task_subnumber": (
        "task_subnumber",
        "subnumber",
        "подномер",
        "подномер_задания",
        "подномер задания",
    ),
    "task_code": ("task_code", "code", "код", "код_задания", "код задания"),
    "official_code": ("official_code", "официальный_код", "официальный код"),
    "max_score": ("max_score", "max_primary_score", "максимальный_балл", "максимальный балл", "балл"),
    "checked_skill": (
        "checked_skill",
        "skill",
        "проверяемое_умение",
        "проверяемое умение",
        "умение",
    ),
    "fgos_result": (
        "fgos_result",
        "planned_result",
        "planned",
        "предметный_результат_фгос",
        "предметный результат фгос",
        "планируемый_результат",
        "планируемый результат",
        "фгос",
    ),
    "program_section": (
        "program_section",
        "section",
        "раздел_программы",
        "раздел программы",
        "раздел",
    ),
    "topic": ("topic", "тема"),
    "topic_subsection": (
        "topic_subsection",
        "subtopic",
        "подтема",
        "подраздел_темы",
        "подраздел темы",
    ),
    "difficulty": (
        "difficulty",
        "complexity",
        "уровень_сложности",
        "уровень сложности",
        "сложность",
    ),
    "task_type": ("task_type", "тип_задания", "тип задания", "тип"),
    "short_description": (
        "short_description",
        "description",
        "описание",
        "краткое_описание",
        "краткое описание",
    ),
    "normative_source": (
        "normative_source",
        "source",
        "источник",
        "нормативный_источник",
        "нормативный источник",
        "нормативный_документ",
        "нормативный документ",
    ),
}


@dataclass
class CatalogImportStats:
    created: int = 0
    updated: int = 0
    skipped: int = 0
    errors: int = 0
    messages: list[str] = field(default_factory=list)

    @property
    def status(self) -> str:
        if self.errors and (self.created or self.updated):
            return VprTaskCatalogImportStatus.PARTIAL
        if self.errors and not self.created and not self.updated:
            return VprTaskCatalogImportStatus.FAILED
        return VprTaskCatalogImportStatus.SUCCESS


def _norm_header(value: Any) -> str:
    text = str(value or "").strip().lower().replace("ё", "е")
    text = text.replace("-", "_").replace(" ", "_")
    while "__" in text:
        text = text.replace("__", "_")
    return text


def _map_headers(headers: Iterable[Any]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    normalized = [_norm_header(h) for h in headers]
    for idx, header in enumerate(normalized):
        for field_name, aliases in FIELD_ALIASES.items():
            alias_norms = {_norm_header(a) for a in aliases}
            if header in alias_norms and field_name not in mapping:
                mapping[field_name] = idx
    return mapping


def _as_int(value: Any, default: int | None = None) -> int | None:
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    text = str(value).strip().replace(",", ".")
    try:
        return int(float(text))
    except ValueError:
        return default


def _as_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _row_to_payload(row: dict[str, Any]) -> dict[str, Any] | None:
    subject = _as_str(row.get("subject"))
    year = _as_int(row.get("academic_year"))
    parallel = _as_int(row.get("parallel"))
    if not subject or year is None or parallel is None:
        return None

    task_code = _as_str(row.get("task_code"))
    task_number = _as_int(row.get("task_number"))
    task_subnumber = _as_str(row.get("task_subnumber"))

    if task_code and task_number is None:
        parsed_num, parsed_sub, canon = parse_task_code(task_code)
        task_number = parsed_num
        if not task_subnumber:
            task_subnumber = parsed_sub
        task_code = canon or task_code
    if task_number is None:
        return None

    if not task_code:
        if task_subnumber.upper().startswith("К") or task_subnumber.upper().startswith("K"):
            raw_code = f"{task_number}{task_subnumber}"
        elif task_subnumber:
            raw_code = f"{task_number}.{task_subnumber}"
        else:
            raw_code = str(task_number)
        _, _, task_code = parse_task_code(raw_code)

    payload = {
        "academic_year": year,
        "subject": subject,
        "parallel": parallel,
        "task_number": task_number,
        "task_subnumber": task_subnumber,
        "task_code": task_code,
        "official_code": _as_str(row.get("official_code")),
        "max_score": _as_int(row.get("max_score"), 0) or 0,
        "checked_skill": _as_str(row.get("checked_skill")),
        "fgos_result": _as_str(row.get("fgos_result")),
        "program_section": _as_str(row.get("program_section")),
        "topic": _as_str(row.get("topic")),
        "topic_subsection": _as_str(row.get("topic_subsection")),
        "difficulty": _as_str(row.get("difficulty")),
        "task_type": _as_str(row.get("task_type")),
        "short_description": _as_str(row.get("short_description")),
        "normative_source": _as_str(row.get("normative_source")),
        "is_active": True,
    }
    extra = row.get("extra")
    if isinstance(extra, dict):
        payload["extra"] = extra
    return payload


def _upsert_payload(payload: dict[str, Any], stats: CatalogImportStats) -> None:
    lookup = {
        "academic_year": payload["academic_year"],
        "subject": payload["subject"],
        "parallel": payload["parallel"],
        "task_number": payload["task_number"],
        "task_subnumber": payload.get("task_subnumber") or "",
    }
    defaults = {k: v for k, v in payload.items() if k not in lookup}
    obj, created = VprTaskCatalogEntry.objects.update_or_create(defaults=defaults, **lookup)
    if created:
        stats.created += 1
    else:
        stats.updated += 1
        _ = obj  # silence lint


def _iter_dict_rows_from_mapping(headers: list[Any], rows: Iterable[Iterable[Any]]) -> list[dict[str, Any]]:
    mapping = _map_headers(headers)
    if "subject" not in mapping or "academic_year" not in mapping or "parallel" not in mapping:
        raise VprCatalogImportError(
            "Не найдены обязательные колонки: предмет, учебный год, класс."
        )
    result: list[dict[str, Any]] = []
    for raw in rows:
        values = list(raw)
        item: dict[str, Any] = {}
        for field_name, idx in mapping.items():
            item[field_name] = values[idx] if idx < len(values) else None
        result.append(item)
    return result


def load_rows_from_json(path: Path) -> list[dict[str, Any]]:
    data = json.loads(path.read_text(encoding="utf-8-sig"))
    if isinstance(data, dict):
        data = data.get("items") or data.get("tasks") or data.get("entries") or []
    if not isinstance(data, list):
        raise VprCatalogImportError("JSON должен содержать список заданий или ключ items/tasks.")
    rows: list[dict[str, Any]] = []
    for item in data:
        if not isinstance(item, dict):
            continue
        normalized: dict[str, Any] = {}
        for field_name, aliases in FIELD_ALIASES.items():
            alias_norms = {_norm_header(a) for a in aliases}
            for key, value in item.items():
                if _norm_header(key) in alias_norms:
                    normalized[field_name] = value
                    break
        if isinstance(item.get("extra"), dict):
            normalized["extra"] = item["extra"]
        rows.append(normalized)
    return rows


def load_rows_from_csv(path: Path) -> list[dict[str, Any]]:
    text = path.read_text(encoding="utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    rows_list = list(reader)
    if not rows_list:
        raise VprCatalogImportError("CSV-файл пуст.")
    return _iter_dict_rows_from_mapping(rows_list[0], rows_list[1:])


def load_rows_from_excel(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = list(next(rows_iter))
        except StopIteration as exc:
            raise VprCatalogImportError("Excel-файл пуст.") from exc
        return _iter_dict_rows_from_mapping(headers, rows_iter)
    finally:
        wb.close()


def detect_format(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".json":
        return "json"
    if suffix == ".csv":
        return "csv"
    if suffix in {".xlsx", ".xlsm"}:
        return "xlsx"
    raise VprCatalogImportError(f"Неподдерживаемый формат: {suffix or 'без расширения'}")


def import_catalog_rows(rows: list[dict[str, Any]]) -> CatalogImportStats:
    stats = CatalogImportStats()
    for idx, row in enumerate(rows, start=1):
        try:
            payload = _row_to_payload(row)
            if not payload:
                stats.skipped += 1
                stats.messages.append(f"Строка {idx}: недостаточно данных, пропущена.")
                continue
            _upsert_payload(payload, stats)
        except Exception as exc:  # noqa: BLE001
            stats.errors += 1
            stats.messages.append(f"Строка {idx}: {exc}")
            logger.warning("VPR catalog row import error: %s", exc)
    return stats


def import_catalog_file(
    path: Path | str,
    *,
    user=None,
    store_file: bool = False,
    uploaded_file=None,
) -> tuple[VprTaskCatalogImport, CatalogImportStats]:
    file_path = Path(path)
    if not file_path.exists():
        raise VprCatalogImportError(f"Файл не найден: {file_path}")

    source_format = detect_format(file_path)
    if source_format == "json":
        rows = load_rows_from_json(file_path)
    elif source_format == "csv":
        rows = load_rows_from_csv(file_path)
    else:
        rows = load_rows_from_excel(file_path)

    stats = import_catalog_rows(rows)
    record = VprTaskCatalogImport(
        uploaded_by=user if getattr(user, "is_authenticated", False) else None,
        original_filename=file_path.name,
        source_format=source_format,
        status=stats.status,
        created_count=stats.created,
        updated_count=stats.updated,
        skipped_count=stats.skipped,
        error_count=stats.errors,
        message=(
            f"Создано: {stats.created}, обновлено: {stats.updated}, "
            f"пропущено: {stats.skipped}, ошибок: {stats.errors}."
        ),
        details={"messages": stats.messages[:100]},
    )
    if store_file and uploaded_file is not None:
        record.file = uploaded_file
    record.save()
    return record, stats


def discover_catalog_data_files(root: Path | str | None = None) -> list[Path]:
    from apps.vpr.catalog.loader import discover_catalog_json_files

    return discover_catalog_json_files(root)


def import_catalog_data_tree(
    root: Path | str | None = None,
    *,
    user=None,
) -> tuple[VprTaskCatalogImport, CatalogImportStats]:
    """Импорт всех JSON из apps/vpr/catalog/data (или указанного каталога)."""
    files = discover_catalog_data_files(root)
    if not files:
        raise VprCatalogImportError("JSON-файлы справочника не найдены.")

    totals = CatalogImportStats()
    file_messages: list[str] = []
    for path in files:
        rows = load_rows_from_json(path)
        stats = import_catalog_rows(rows)
        totals.created += stats.created
        totals.updated += stats.updated
        totals.skipped += stats.skipped
        totals.errors += stats.errors
        totals.messages.extend(stats.messages)
        file_messages.append(
            f"{path.name}: +{stats.created}/~{stats.updated}/skip {stats.skipped}/err {stats.errors}"
        )

    record = VprTaskCatalogImport(
        uploaded_by=user if getattr(user, "is_authenticated", False) else None,
        original_filename=f"catalog_data_tree ({len(files)} files)",
        source_format="json",
        status=totals.status,
        created_count=totals.created,
        updated_count=totals.updated,
        skipped_count=totals.skipped,
        error_count=totals.errors,
        message=(
            f"Создано: {totals.created}, обновлено: {totals.updated}, "
            f"пропущено: {totals.skipped}, ошибок: {totals.errors}. "
            f"Файлов: {len(files)}."
        ),
        details={"files": file_messages, "messages": totals.messages[:100]},
    )
    record.save()
    return record, totals


def import_catalog_path(
    path: Path | str | None = None,
    *,
    user=None,
) -> tuple[VprTaskCatalogImport, CatalogImportStats]:
    """
    Импорт файла или каталога.
    Без аргумента — apps/vpr/catalog/data.
    """
    if path is None:
        return import_catalog_data_tree(user=user)
    target = Path(path)
    if target.is_dir():
        return import_catalog_data_tree(target, user=user)
    return import_catalog_file(target, user=user)
