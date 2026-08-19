"""Парсер официальной выгрузки «Ф1 Индивидуальные результаты» ВПР."""

from __future__ import annotations

import logging
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from apps.vpr.exceptions import VprParseError, VprValidationError
from apps.vpr.parsers.base import BaseVprParser
from apps.vpr.parsers.dto import (
    VprParseResult,
    VprStudentRow,
    VprTaskMeta,
    VprTaskScoreData,
)

logger = logging.getLogger(__name__)

TITLE_RE = re.compile(
    r"ВПР\s+(?P<year>\d{4})\s+(?P<subject>.+?)\s+(?P<parallel>\d+)\s*класс",
    re.IGNORECASE,
)
TASK_HEADER_RE = re.compile(
    r"^(?P<code>\d+(?:[.,]\d+)?(?:[КK]\d+)?)\s*\((?P<max>\d+)\s*б\)$",
    re.IGNORECASE,
)
ORG_CODE_RE = re.compile(r"\(?(edu\d+)\)?", re.IGNORECASE)
PARTICIPANTS_RE = re.compile(r"Кол-во\s+участников\s*:\s*(\d+)", re.IGNORECASE)

REQUIRED_HEADERS = {
    "оо": "organization_code",
    "муниципалитет": "municipality",
    "название оо": "organization_name",
    "код участника": "participant_code",
    "класс": "class_group",
    "вариант": "variant",
    "первичный балл": "primary_score",
    "отметка": "mark_vpr",
    "отметка по журналу": "mark_journal",
}


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize_header(value: Any) -> str:
    text = _cell_str(value).lower().replace("ё", "е")
    text = re.sub(r"\s+", " ", text)
    return text


def _parse_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    text = _cell_str(value).replace(",", ".")
    if not text or text.upper() in {"X", "Х", "-", "—", "НЕТ ОТМЕТКИ"}:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def _parse_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = _cell_str(value).replace(",", ".")
    if not text or text.upper() in {"X", "Х", "-", "—"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _parse_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _cell_str(value)
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _extract_org_code(text: str) -> str:
    match = ORG_CODE_RE.search(text or "")
    return match.group(1).lower() if match else ""


class F1IndividualResultsParser(BaseVprParser):
    """
    Шаблон ФИС ОКО / ВПР:
    лист «… Индивидуальные результаты»
    заголовок «ВПР YYYY Предмет N класс»
    таблица с колонками ОО, Код участника, задания, Первичный балл, Отметка, Отметка по журналу.
    """

    template_key = "f1_individual"
    display_name = "Ф1 Индивидуальные результаты"

    def can_parse(self, path: Path) -> bool:
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
        except Exception:
            return False
        try:
            for name in wb.sheetnames:
                if "индивидуальн" in name.lower():
                    return True
                ws = wb[name]
                for row in ws.iter_rows(min_row=1, max_row=12, max_col=8, values_only=True):
                    for cell in row:
                        text = _cell_str(cell)
                        if TITLE_RE.search(text) or text.lower() == "индивидуальные результаты":
                            return True
            return False
        finally:
            wb.close()

    def parse(self, path: Path) -> VprParseResult:
        try:
            wb = load_workbook(path, data_only=True)
        except Exception as exc:
            raise VprValidationError(f"Не удалось прочитать файл Excel: {exc}") from exc

        try:
            sheet_name, ws = self._pick_sheet(wb)
            meta = self._read_meta(ws)
            header_row, columns, tasks = self._read_header(ws)
            students = self._read_students(ws, header_row, columns, tasks)
            if not students:
                raise VprValidationError(
                    "В файле не найдены учащиеся с результатами.",
                    details=["Проверьте, что загружена заполненная форма «Индивидуальные результаты»."],
                )

            first = students[0]
            organization_code = meta.get("organization_code") or first.organization_code
            organization_name = meta.get("organization_name") or first.organization_name
            # Если в метаданных остался только код edu — берём полное название из строк.
            if organization_name and organization_name.lower() == organization_code.lower():
                organization_name = first.organization_name or organization_name
            municipality = meta.get("municipality") or first.municipality

            warnings: list[str] = []
            declared = meta.get("participants_declared")
            if declared is not None and declared != len(students):
                warnings.append(
                    f"Заявлено участников: {declared}, фактически в таблице: {len(students)}."
                )

            result = VprParseResult(
                template_key=self.template_key,
                sheet_name=sheet_name,
                source_title=meta["source_title"],
                subject=meta["subject"],
                parallel=meta["parallel"],
                academic_year=meta["academic_year"],
                exam_date=meta.get("exam_date"),
                max_primary_score=int(meta.get("max_primary_score") or 0),
                organization_code=organization_code,
                organization_name=organization_name,
                municipality=municipality,
                participants_declared=declared,
                tasks=tasks,
                students=students,
                warnings=warnings,
            )
            logger.info(
                "VPR F1 parsed: subject=%s parallel=%s students=%s tasks=%s file=%s",
                result.subject,
                result.parallel,
                result.participants_count,
                result.tasks_count,
                path.name,
            )
            return result
        finally:
            wb.close()

    def _pick_sheet(self, wb):
        preferred = []
        for name in wb.sheetnames:
            low = name.lower()
            if "индивидуальн" in low:
                preferred.append(name)
        candidates = preferred or list(wb.sheetnames)
        for name in candidates:
            ws = wb[name]
            # look for title or header markers
            for row in ws.iter_rows(min_row=1, max_row=15, max_col=10, values_only=True):
                joined = " ".join(_cell_str(c) for c in row if c is not None)
                if TITLE_RE.search(joined) or "код участника" in joined.lower():
                    return name, ws
        raise VprValidationError(
            "Не найден лист с индивидуальными результатами ВПР.",
            details=[f"Доступные листы: {', '.join(wb.sheetnames)}"],
        )

    def _read_meta(self, ws) -> dict[str, Any]:
        source_title = ""
        subject = ""
        parallel = None
        academic_year = None
        exam_date = None
        max_primary_score = 0
        organization_name = ""
        organization_code = ""
        municipality = ""
        participants_declared = None
        details: list[str] = []

        for r in range(1, 12):
            label = _normalize_header(ws.cell(r, 1).value)
            value = ws.cell(r, 2).value
            row_text = _cell_str(ws.cell(r, 1).value)
            if not source_title:
                match = TITLE_RE.search(row_text)
                if match:
                    source_title = row_text
                    academic_year = int(match.group("year"))
                    subject = match.group("subject").strip()
                    parallel = int(match.group("parallel"))
            if label.startswith("предмет"):
                subject = _cell_str(value) or subject
            elif "максимальный первичный балл" in label:
                max_primary_score = _parse_int(value) or 0
            elif label.startswith("дата"):
                exam_date = _parse_date(value)
            elif "кол-во участников" in row_text.lower():
                organization_name = row_text.split("|")[0].strip()
                organization_code = _extract_org_code(row_text) or organization_code
                m = PARTICIPANTS_RE.search(row_text)
                if m:
                    participants_declared = int(m.group(1))
            elif not organization_name and ORG_CODE_RE.search(row_text) and len(row_text) > 20:
                # Строка-заголовок ОО без явного «Кол-во участников».
                organization_name = row_text.split("|")[0].strip()
                organization_code = _extract_org_code(row_text) or organization_code
                m = PARTICIPANTS_RE.search(row_text)
                if m:
                    participants_declared = int(m.group(1))

        if not source_title:
            details.append("Не найден заголовок вида «ВПР YYYY Предмет N класс».")
        if not subject:
            details.append("Не найден предмет.")
        if parallel is None:
            details.append("Не найден класс (параллель).")
        if academic_year is None:
            details.append("Не найден учебный год.")
        if details:
            raise VprValidationError("Не удалось извлечь метаданные протокола ВПР.", details=details)

        return {
            "source_title": source_title,
            "subject": subject,
            "parallel": parallel,
            "academic_year": academic_year,
            "exam_date": exam_date,
            "max_primary_score": max_primary_score,
            "organization_name": organization_name,
            "organization_code": organization_code,
            "municipality": municipality,
            "participants_declared": participants_declared,
        }

    def _read_header(self, ws) -> tuple[int, dict[str, int], list[VprTaskMeta]]:
        header_row = None
        header_map: dict[str, int] = {}
        for r in range(1, 30):
            values = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            norms = [_normalize_header(v) for v in values]
            if "код участника" in norms and "первичный балл" in norms:
                header_row = r
                for idx, norm in enumerate(norms, start=1):
                    if norm:
                        header_map[norm] = idx
                break
        if header_row is None:
            raise VprValidationError(
                "Не найдена строка заголовков таблицы участников.",
                details=["Ожидаются колонки «Код участника», «Первичный балл», «Отметка»."],
            )

        missing = [label for label in REQUIRED_HEADERS if label not in header_map]
        # «ОО» may appear as exact key
        if "оо" not in header_map and "оо" in missing:
            # tolerate missing ОО if organization present in meta row
            missing = [m for m in missing if m != "оо"]
        critical = [m for m in missing if m in {"код участника", "первичный балл", "отметка"}]
        if critical:
            raise VprValidationError(
                "В таблице отсутствуют обязательные заголовки.",
                details=[f"Нет колонок: {', '.join(critical)}"],
            )

        # Difficulty row is usually header_row - 1
        difficulty_by_col: dict[int, str] = {}
        if header_row > 1:
            for c in range(1, ws.max_column + 1):
                diff = _cell_str(ws.cell(header_row - 1, c).value)
                if diff in {"Б", "П", "б", "п"}:
                    difficulty_by_col[c] = diff.upper()

        tasks: list[VprTaskMeta] = []
        columns: dict[str, int] = {}
        for label, field_name in REQUIRED_HEADERS.items():
            if label in header_map:
                columns[field_name] = header_map[label]

        position = 0
        for c in range(1, ws.max_column + 1):
            title = _cell_str(ws.cell(header_row, c).value)
            match = TASK_HEADER_RE.match(title.replace(" ", ""))
            if not match:
                # try with spaces: "1 (2б)"
                match = TASK_HEADER_RE.match(title)
            if not match:
                continue
            position += 1
            code = match.group("code").replace(",", ".")
            max_score = int(match.group("max"))
            tasks.append(
                VprTaskMeta(
                    position=position,
                    code=code,
                    title=title,
                    max_score=max_score,
                    difficulty=difficulty_by_col.get(c, ""),
                )
            )
            columns[f"task:{code}"] = c

        if not tasks:
            raise VprValidationError(
                "Не найдены колонки заданий ВПР.",
                details=["Ожидаются заголовки вида «1 (2б)», «9.1 (1б)»."],
            )

        return header_row, columns, tasks

    def _read_students(
        self,
        ws,
        header_row: int,
        columns: dict[str, int],
        tasks: list[VprTaskMeta],
    ) -> list[VprStudentRow]:
        students: list[VprStudentRow] = []
        code_col = columns["participant_code"]
        for r in range(header_row + 1, ws.max_row + 1):
            participant_code = _cell_str(ws.cell(r, code_col).value)
            if not participant_code:
                # stop on long empty streak? keep scanning — skip empty
                continue
            # skip section headers like school name rows without numeric code
            if not re.search(r"\d", participant_code):
                continue

            task_scores: list[VprTaskScoreData] = []
            for task in tasks:
                col = columns.get(f"task:{task.code}")
                raw = ws.cell(r, col).value if col else None
                raw_text = _cell_str(raw)
                score = _parse_float(raw)
                task_scores.append(
                    VprTaskScoreData(
                        task_code=task.code,
                        raw_value=raw_text,
                        score=score,
                        max_score=task.max_score,
                    )
                )

            def col_val(key: str) -> Any:
                col = columns.get(key)
                return ws.cell(r, col).value if col else None

            org_name = _cell_str(col_val("organization_name"))
            org_code = _cell_str(col_val("organization_code")) or _extract_org_code(org_name)

            students.append(
                VprStudentRow(
                    participant_code=participant_code,
                    organization_code=org_code,
                    organization_name=org_name,
                    municipality=_cell_str(col_val("municipality")),
                    class_group=_cell_str(col_val("class_group")),
                    variant=_cell_str(col_val("variant")),
                    gender=_cell_str(col_val("gender")),
                    full_name=_cell_str(col_val("full_name")),
                    primary_score=_parse_float(col_val("primary_score")),
                    mark_vpr=_parse_int(col_val("mark_vpr")),
                    mark_journal=_parse_int(col_val("mark_journal")),
                    source_row=r,
                    task_scores=task_scores,
                )
            )
        return students
