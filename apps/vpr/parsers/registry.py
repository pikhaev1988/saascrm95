"""Реестр парсеров ВПР — расширяется новыми шаблонами без изменения старых."""

from __future__ import annotations

import logging
from pathlib import Path

from apps.vpr.exceptions import VprValidationError
from apps.vpr.parsers.base import BaseVprParser
from apps.vpr.parsers.dto import VprParseResult
from apps.vpr.parsers.f1_individual import F1IndividualResultsParser

logger = logging.getLogger(__name__)

_PARSERS: list[BaseVprParser] = [
    F1IndividualResultsParser(),
]


def register_parser(parser: BaseVprParser) -> None:
    """Зарегистрировать дополнительный шаблон ВПР."""
    _PARSERS.append(parser)


def list_parsers() -> list[BaseVprParser]:
    return list(_PARSERS)


def detect_and_parse(path: Path | str) -> VprParseResult:
    """
    Автоматически выбрать подходящий парсер и разобрать файл.
    """
    file_path = Path(path)
    if not file_path.exists():
        raise VprValidationError(f"Файл не найден: {file_path}")

    suffix = file_path.suffix.lower()
    if suffix not in {".xlsx", ".xlsm"}:
        raise VprValidationError(
            "Поддерживаются только файлы Excel (.xlsx).",
            details=[f"Получено расширение: {suffix or 'без расширения'}"],
        )

    # Пустая «Форма сбора результатов» — частая ошибка пользователя.
    try:
        from openpyxl import load_workbook

        wb = load_workbook(file_path, read_only=True, data_only=True)
        try:
            names = [n.lower() for n in wb.sheetnames]
            if "протокол" in names and not any("индивидуальн" in n for n in names):
                raise VprValidationError(
                    "Загружена пустая «Форма сбора результатов», а не протокол с итогами.",
                    details=[
                        "Нужна выгрузка «Ф1 Индивидуальные результаты» "
                        "(лист с колонками «Код участника», «Первичный балл», «Отметка»).",
                    ],
                )
        finally:
            wb.close()
    except VprValidationError:
        raise
    except Exception:
        pass

    matched: list[BaseVprParser] = []
    for parser in _PARSERS:
        try:
            if parser.can_parse(file_path):
                matched.append(parser)
        except Exception as exc:  # noqa: BLE001 — изоляция шаблонов
            logger.warning("VPR can_parse failed for %s: %s", parser.template_key, exc)

    if not matched:
        if _PARSERS:
            return _PARSERS[0].parse(file_path)
        raise VprValidationError("Не зарегистрировано ни одного парсера ВПР.")

    parser = matched[0]
    logger.info("VPR template selected: %s for %s", parser.template_key, file_path.name)
    return parser.parse(file_path)


class VprExcelParser:
    """
    Публичный фасад парсера ВПР.
    Точка входа для сервисов и views.
    """

    def parse(self, path: Path | str) -> VprParseResult:
        return detect_and_parse(path)

    def can_parse(self, path: Path | str) -> bool:
        file_path = Path(path)
        return any(p.can_parse(file_path) for p in _PARSERS)
