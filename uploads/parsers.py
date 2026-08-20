import datetime as dt
import csv
import os
import re
from dataclasses import dataclass
from typing import Iterable

from openpyxl import load_workbook
import xlrd

from exams.models import Exam, ExamResult, Student, TaskResult
from organizations.models import School
from users.task_topics import parse_long_answer_mask


EXAM_HEADER_PATTERN = re.compile(r"(?P<code>\d+)\s*-\s*(?P<subject>.+)\s+(?P<date>\d{4}\.\d{2}\.\d{2})")
OGE_FILE_PATTERN = re.compile(r"(?P<code>\d{2})_(?P<date>\d{4}\.\d{2}\.\d{2})")
OGE_EXAM_CELL_PATTERN = re.compile(
    r"(?P<subject>.+?)\s*\(\s*(?P<date>\d{4}\.\d{2}\.\d{2})\s*\)"
)

OGE_SUBJECT_CODE_MAP = {
    "01": "Русский язык",
    "02": "Математика",
    "03": "Физика",
    "04": "Химия",
    "05": "Информатика",
    "06": "Биология",
    "07": "История",
    "08": "География",
    "09": "Обществознание",
    "10": "Литература",
    "11": "Английский язык",
    "12": "Немецкий язык",
    "13": "Французский язык",
    "14": "Испанский язык",
    "18": "Родной язык",
    "94": "Резерв",
}

OGE_SUBJECT_NAME_TO_CODE = {name.lower(): code for code, name in OGE_SUBJECT_CODE_MAP.items()}


@dataclass
class ParsedRow:
    school_code: str
    student_id: str
    full_name: str
    grade: str
    short_answer_tasks: str
    long_answer_tasks: str
    primary_score: float
    score: float
    total_score: float
    source_row: dict


@dataclass
class ParseStats:
    exams_processed: int = 0
    results_imported: int = 0
    skipped_other_school: int = 0
    skipped_unknown_school: int = 0


def _normalize_school_code(code) -> str:
    text = _safe_str(code)
    if not text:
        return ""
    try:
        number = float(text.replace(",", "."))
        if number == int(number):
            return str(int(number))
    except ValueError:
        pass
    return text


def parse_exam_header(raw_header: str):
    match = EXAM_HEADER_PATTERN.search((raw_header or "").strip())
    if not match:
        raise ValueError(f"Некорректный формат экзамена: {raw_header}")
    exam_date = dt.datetime.strptime(match.group("date"), "%Y.%m.%d").date()
    return match.group("code"), match.group("subject").strip(), exam_date


def _oge_subject_to_code(subject: str) -> str:
    normalized = (subject or "").strip().lower()
    if not normalized:
        return "00"
    if normalized in OGE_SUBJECT_NAME_TO_CODE:
        return OGE_SUBJECT_NAME_TO_CODE[normalized]
    for name, code in OGE_SUBJECT_NAME_TO_CODE.items():
        if normalized.startswith(name) or name.startswith(normalized):
            return code
    return "00"


def parse_oge_exam_cell(raw_cell: str):
    """Parse cells like 'Математика(2026.06.02)' from appeal-results protocols."""
    match = OGE_EXAM_CELL_PATTERN.search((raw_cell or "").strip())
    if not match:
        raise ValueError(f"Некорректный формат экзамена в ячейке: {raw_cell}")
    subject = match.group("subject").strip()
    exam_date = dt.datetime.strptime(match.group("date"), "%Y.%m.%d").date()
    code = _oge_subject_to_code(subject)
    return code, subject, exam_date


def _find_exam_header_in_row(row):
    for value in row:
        if value is None:
            continue
        text = str(value).strip()
        if not text:
            continue
        if EXAM_HEADER_PATTERN.search(text):
            return text
    return None


def _safe_float(value):
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", ".")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        match = re.search(r"-?\d+(?:\.\d+)?", text)
        return float(match.group(0)) if match else 0.0


def _iter_rows(file_path, start_row):
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=start_row, values_only=True):
        if not row or not row[0]:
            continue
        yield ParsedRow(
            school_code=str(row[0]).strip(),
            student_id=str(row[1]).strip(),
            full_name=str(row[2]).strip(),
            grade=str(row[3] or "").strip(),
            short_answer_tasks=str(row[4] or "").strip(),
            long_answer_tasks=str(row[5] or "").strip(),
            primary_score=_safe_float(row[6]),
            score=_safe_float(row[7]),
            total_score=_safe_float(row[7]),
            source_row={"raw_values": ["" if v is None else str(v) for v in row]},
        )


def _normalize_header_cell(text) -> str:
    return _safe_str(text).lower().replace("ё", "е")


def _find_header_index(headers, *patterns: str) -> int | None:
    for idx, header in enumerate(headers):
        normalized = _normalize_header_cell(header)
        if not normalized:
            continue
        for pattern in patterns:
            if pattern in normalized:
                return idx
    return None


def _build_ege_column_map(header_row) -> dict:
    headers = list(header_row)
    msu_idx = _find_header_index(headers, "код мсу")
    oo_idx = _find_header_index(headers, "код оо")
    participant_idx = _find_header_index(headers, "код участника")

    # Код ОО — код школы; Код МСУ — муниципалитет (не подменяет код школы).
    if oo_idx is not None:
        school_idx = oo_idx
    else:
        school_idx = msu_idx

    total_idx = _find_header_index(
        headers,
        "тестовый балл",
        "итоговый балл",
        "итоговая оценка",
        "оценка",
    )
    primary_idx = _find_header_index(headers, "первичный балл")

    return {
        "order": _find_header_index(headers, "№") or 0,
        "school_code": school_idx,
        "participant_code": participant_idx,
        "grade": _find_header_index(headers, "класс"),
        "surname": _find_header_index(headers, "фамилия"),
        "name": _find_header_index(headers, "имя"),
        "patronymic": _find_header_index(headers, "отчество"),
        "short_tasks": _find_header_index(
            headers,
            "кратким ответом",
            "краткой форме",
        ),
        "long_tasks": _find_header_index(
            headers,
            "развернутым ответом",
            "развёрнутым ответом",
        ),
        "primary_score": primary_idx,
        "total_score": total_idx,
    }


def _default_ege_column_map(layout: str) -> dict:
    if layout == "regional_sparse":
        return {
            "order": 0,
            "school_code": 4,
            "participant_code": None,
            "grade": 6,
            "surname": 12,
            "name": 13,
            "patronymic": 14,
            "short_tasks": 16,
            "long_tasks": 18,
            "primary_score": 20,
            "total_score": 22,
        }
    return {
        "order": 0,
        "school_code": 2,
        "participant_code": 10,
        "grade": 3,
        "surname": 6,
        "name": 7,
        "patronymic": 8,
        "short_tasks": 12,
        "long_tasks": 13,
        "primary_score": 14,
        "total_score": 15,
    }


def _ege_cell(row, index: int | None):
    if index is None or index >= len(row):
        return None
    return row[index]


def _is_ege_column_header_row(row) -> bool:
    headers = [_normalize_header_cell(value) for value in row]
    if any(header == "код оо" for header in headers):
        return True
    if any("код мсу" in header for header in headers if header):
        return True
    if any("код участника" in header for header in headers if header):
        return True
    return False


def _detect_ege_layout_from_header(row) -> str:
    headers = [_safe_str(value).lower() for value in row]
    header_joined = " ".join(header for header in headers if header)
    is_sparse = len(row) > 3 and row[1] is None and row[3] is None

    if is_sparse and any(
        marker in header_joined
        for marker in (
            "код мсу",
            "тестовый балл",
            "задания с кратким ответом",
            "задания с развёрнутым ответом",
            "задания с развернутым ответом",
        )
    ):
        return "regional_sparse"

    for idx, header in enumerate(headers):
        if idx == 10 and header and "код участника" in header:
            return "compact"
    if is_sparse and any(header == "код оо" for header in headers):
        return "regional_sparse"
    return "compact"


def _infer_ege_layout_from_data_row(row) -> str | None:
    if len(row) < 15 or not isinstance(row[0], (int, float)):
        return None
    regional_surname = _safe_str(row[12] if len(row) > 12 else "")
    compact_surname = _safe_str(row[6] if len(row) > 6 else "")
    if regional_surname and any(char.isalpha() for char in regional_surname):
        return "regional_sparse"
    if compact_surname and any(char.isalpha() for char in compact_surname):
        return "compact"
    return None


def _parse_ege_row(row, column_map: dict) -> ParsedRow | None:
    order_idx = column_map.get("order", 0)
    if not isinstance(_ege_cell(row, order_idx), (int, float)):
        return None

    school_code = _safe_str(_ege_cell(row, column_map.get("school_code")))
    if not school_code:
        return None
    if school_code.lower() in {"код оо", "код мсу"}:
        return None

    surname = _safe_str(_ege_cell(row, column_map.get("surname")))
    name = _safe_str(_ege_cell(row, column_map.get("name")))
    patronymic = _safe_str(_ege_cell(row, column_map.get("patronymic")))
    participant_code = _safe_str(_ege_cell(row, column_map.get("participant_code")))
    if not participant_code:
        participant_code = " ".join(part for part in (surname, name, patronymic) if part)
    if not participant_code:
        participant_code = _safe_str(_ege_cell(row, order_idx))
    if not participant_code:
        return None
    task_mask = _safe_str(_ege_cell(row, column_map.get("short_tasks")))
    long_answer = _safe_str(_ege_cell(row, column_map.get("long_tasks")))
    primary_score = _ege_cell(row, column_map.get("primary_score"))
    total_score = _ege_cell(row, column_map.get("total_score"))
    if total_score is None:
        total_score = primary_score

    grade = _safe_str(_ege_cell(row, column_map.get("grade"))) or "ЕГЭ"
    full_name = " ".join(part for part in (surname, name, patronymic) if part)
    return ParsedRow(
        school_code=school_code,
        student_id=participant_code,
        full_name=full_name or participant_code,
        grade=grade,
        short_answer_tasks=task_mask,
        long_answer_tasks=long_answer,
        primary_score=_safe_float(primary_score),
        score=_safe_float(total_score),
        total_score=_safe_float(total_score),
        source_row={
            "raw_values": ["" if v is None else str(v) for v in row],
            "column_map": column_map,
        },
    )


def _iter_ege_rows(file_path, start_row=3) -> Iterable[ParsedRow]:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet = workbook.active
    header_row = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True), ())
    column_map = _build_ege_column_map(header_row)
    headers = ["" if v is None else str(v) for v in header_row]
    for row in sheet.iter_rows(min_row=start_row, values_only=True):
        parsed = _parse_ege_row(row, column_map)
        if not parsed:
            continue
        yield ParsedRow(
            school_code=parsed.school_code,
            student_id=parsed.student_id,
            full_name=parsed.full_name,
            grade=parsed.grade,
            short_answer_tasks=parsed.short_answer_tasks,
            long_answer_tasks=parsed.long_answer_tasks,
            primary_score=parsed.primary_score,
            score=parsed.score,
            total_score=parsed.total_score,
            source_row={
                "raw_values": parsed.source_row.get("raw_values", []),
                "headers": headers,
                "column_map": column_map,
            },
        )


def _iter_ege_exam_blocks(file_path, start_row=1):
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet = workbook.active
    current_exam_key = None
    current_layout = None
    current_column_map = None
    rows_by_exam = {}
    for row_index, row in enumerate(sheet.iter_rows(min_row=start_row, values_only=True), start=start_row):
        header_text = _find_exam_header_in_row(row)
        if header_text:
            code, subject, exam_date = parse_exam_header(header_text)
            current_exam_key = (code, subject, exam_date)
            rows_by_exam.setdefault(current_exam_key, [])
            current_layout = None
            current_column_map = None
            continue
        if not current_exam_key:
            continue
        if _is_ege_column_header_row(row):
            current_layout = _detect_ege_layout_from_header(row)
            current_column_map = _build_ege_column_map(row)
            continue
        layout = current_layout or _infer_ege_layout_from_data_row(row) or "compact"
        if current_layout is None:
            current_layout = layout
        column_map = current_column_map or _default_ege_column_map(layout)
        parsed = _parse_ege_row(row, column_map)
        if not parsed:
            continue
        rows_by_exam[current_exam_key].append(
            ParsedRow(
                school_code=parsed.school_code,
                student_id=parsed.student_id,
                full_name=parsed.full_name,
                grade=parsed.grade,
                short_answer_tasks=parsed.short_answer_tasks,
                long_answer_tasks=parsed.long_answer_tasks,
                primary_score=parsed.primary_score,
                score=parsed.score,
                total_score=parsed.total_score,
                source_row={
                    "row_number": row_index,
                    "layout": layout,
                    "column_map": column_map,
                    "exam_header": f"{current_exam_key[0]} - {current_exam_key[1]} {current_exam_key[2]:%Y.%m.%d}",
                    "raw_values": parsed.source_row.get("raw_values", []),
                },
            )
        )
    return rows_by_exam


def _safe_str(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def _build_oge_student_id(row_data):
    series = _safe_str(row_data.get("DocumentSeriesCode"))
    number = _safe_str(row_data.get("DocumentNumberCode"))
    if series or number:
        return f"{series}{number}"
    return _safe_str(row_data.get("GovernmentCell")) or _safe_str(row_data.get("ColumnName"))


def _oge_cell(row, index: int | None):
    if index is None or index >= len(row):
        return None
    return row[index]


def _build_oge_column_map(header_row) -> dict:
    headers = list(header_row)
    return {
        "order": _find_header_index(headers, "№") or 0,
        "school_code": _find_header_index(headers, "код оо"),
        "class_code": _find_header_index(headers, "класс"),
        "station_code": _find_header_index(headers, "код ппэ"),
        "auditorium_code": _find_header_index(headers, "аудитория"),
        "msu_code": _find_header_index(headers, "код мсу"),
        "government_cell": _find_header_index(headers, "код регистрации", "гос номер"),
        "surname": _find_header_index(headers, "фамилия"),
        "name": _find_header_index(headers, "имя"),
        "patronymic": _find_header_index(headers, "отчество"),
        "document_series": _find_header_index(headers, "серия"),
        "document_number": _find_header_index(headers, "номер"),
        "short_tasks": _find_header_index(headers, "кратким ответом", "краткой форме"),
        "long_tasks": _find_header_index(
            headers,
            "развернутым ответом",
            "развёрнутым ответом",
        ),
        "primary_mark": _find_header_index(headers, "первичный балл"),
        "mark5": _find_header_index(headers, "оценка"),
    }


def _default_oge_column_map(layout: str) -> dict:
    if layout == "regional_sparse":
        return {
            "order": 0,
            "school_code": 1,
            "class_code": 2,
            "station_code": 4,
            "auditorium_code": 6,
            "msu_code": 8,
            "government_cell": None,
            "surname": 10,
            "name": 12,
            "patronymic": 13,
            "document_series": None,
            "document_number": None,
            "short_tasks": 14,
            "long_tasks": 17,
            "primary_mark": 18,
            "mark5": 20,
        }
    if layout == "appeal_results":
        return {
            "order": 0,
            "school_code": 1,
            "class_code": 2,
            "station_code": 3,
            "auditorium_code": 4,
            "government_cell": None,
            "surname": 5,
            "name": 6,
            "patronymic": 7,
            "exam_cell": 8,
            "document_series": None,
            "document_number": None,
            "short_tasks": None,
            "long_tasks": None,
            "primary_mark": 10,
            "mark5": 11,
        }
    return {
        "order": 0,
        "school_code": 1,
        "class_code": 2,
        "station_code": 3,
        "auditorium_code": 4,
        "government_cell": 5,
        "surname": 6,
        "name": 7,
        "patronymic": 8,
        "document_series": 9,
        "document_number": 10,
        "short_tasks": 11,
        "long_tasks": 12,
        "primary_mark": 13,
        "mark5": 14,
    }


def _is_oge_appeal_results_header_row(row) -> bool:
    headers = [_normalize_header_cell(value) for value in row if value is not None]
    has_exam = any(header == "экзамен" for header in headers)
    has_appeal = any("апелляц" in header for header in headers if header)
    return has_exam and has_appeal


def _is_oge_appeal_results_xlsx(file_path) -> bool:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=1, max_row=20, values_only=True):
        if row and _is_oge_appeal_results_header_row(row):
            return True
    return False


def _is_oge_column_header_row(row) -> bool:
    headers = [_normalize_header_cell(value) for value in row]
    if any(header == "schoolcode" for header in headers):
        return True
    if any(header == "код оо" for header in headers):
        return True
    if any(header == "фамилия" for header in headers):
        return True
    return False


def _detect_oge_layout_from_header(row) -> str:
    headers = [_safe_str(value).lower() for value in row]
    header_joined = " ".join(header for header in headers if header)
    # openpyxl uses None for empty cells; xlrd uses "".
    is_sparse = (
        len(row) > 4
        and (row[3] is None or (isinstance(row[3], str) and not row[3].strip()))
        and any(header == "код оо" for header in headers)
    )

    if any(header == "schoolcode" for header in headers):
        return "normalized"
    if is_sparse and ("код оо" in header_joined or "код мсу" in header_joined):
        return "regional_sparse"
    return "legacy_dense"


def _infer_oge_layout_from_data_row(row) -> str | None:
    if len(row) < 15 or not isinstance(row[0], (int, float)):
        return None
    regional_surname = _safe_str(row[10] if len(row) > 10 else "")
    legacy_surname = _safe_str(row[6] if len(row) > 6 else "")
    legacy_is_name = legacy_surname and any(char.isalpha() for char in legacy_surname)
    regional_is_name = regional_surname and any(char.isalpha() for char in regional_surname)
    if regional_is_name and not legacy_is_name:
        return "regional_sparse"
    if legacy_is_name:
        return "legacy_dense"
    return None


def _parse_oge_row(row, column_map: dict) -> ParsedRow | None:
    order_idx = column_map.get("order", 0)
    if not isinstance(_oge_cell(row, order_idx), (int, float)):
        return None

    school_code = _safe_str(_oge_cell(row, column_map.get("school_code")))
    if not school_code or school_code.lower() in {"код оо", "код мсу"}:
        return None

    surname = _safe_str(_oge_cell(row, column_map.get("surname")))
    name = _safe_str(_oge_cell(row, column_map.get("name")))
    patronymic = _safe_str(_oge_cell(row, column_map.get("patronymic")))
    full_name = " ".join(part for part in (surname, name, patronymic) if part)

    series = _safe_str(_oge_cell(row, column_map.get("document_series")))
    number = _safe_str(_oge_cell(row, column_map.get("document_number")))
    student_id = f"{series}{number}" if (series or number) else ""
    if not student_id:
        student_id = full_name
    if not student_id:
        student_id = _safe_str(_oge_cell(row, column_map.get("government_cell")))
    if not student_id:
        student_id = _safe_str(_oge_cell(row, order_idx))
    if not student_id:
        return None

    primary_mark = _safe_float(_oge_cell(row, column_map.get("primary_mark")))
    grade_mark = _safe_float(_oge_cell(row, column_map.get("mark5")))
    final_mark = grade_mark if grade_mark > 0 else primary_mark

    return ParsedRow(
        school_code=school_code,
        student_id=student_id,
        full_name=full_name or student_id,
        grade=_safe_str(_oge_cell(row, column_map.get("class_code"))),
        short_answer_tasks=_safe_str(_oge_cell(row, column_map.get("short_tasks"))),
        long_answer_tasks=_safe_str(_oge_cell(row, column_map.get("long_tasks"))),
        primary_score=primary_mark,
        score=final_mark,
        total_score=final_mark,
        source_row={
            "raw_values": ["" if v is None else str(v) for v in row],
            "column_map": column_map,
        },
    )


def _oge_row_to_parsed(row_data):
    school_code = _safe_str(row_data.get("SchoolCode"))
    student_id = _build_oge_student_id(row_data)
    surname = _safe_str(row_data.get("SurnameCell"))
    name = _safe_str(row_data.get("NameCell"))
    second_name = _safe_str(row_data.get("SecondNameCell"))
    full_name = " ".join(part for part in (surname, name, second_name) if part)
    if not school_code or not student_id:
        return None
    short_tasks = _safe_str(row_data.get("TestResultBCell"))
    long_tasks = _safe_str(row_data.get("TestResultCCell"))
    primary_mark = _safe_float(row_data.get("PrimaryMark"))
    grade_mark = _safe_float(row_data.get("Mark5"))
    final_mark = grade_mark if grade_mark > 0 else primary_mark
    return ParsedRow(
        school_code=school_code,
        student_id=student_id,
        full_name=full_name or student_id,
        grade=_safe_str(row_data.get("ClassCode")),
        short_answer_tasks=short_tasks,
        long_answer_tasks=long_tasks,
        primary_score=primary_mark,
        score=final_mark,
        total_score=final_mark,
        source_row={"raw_values": row_data},
    )


def _iter_oge_csv_rows(file_path):
    with open(file_path, "r", encoding="utf-8-sig", newline="") as source:
        reader = csv.DictReader(source, delimiter=";")
        for row in reader:
            parsed = _oge_row_to_parsed(row)
            if parsed:
                yield parsed


def _iter_oge_xlsx_rows(file_path):
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet = workbook.active
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    headers = [_safe_str(value) for value in header_row]

    # Variant A: normalized xlsx export with explicit English headers.
    if "SchoolCode" in headers and "PrimaryMark" in headers:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = {headers[idx]: row[idx] if idx < len(row) else "" for idx in range(len(headers))}
            parsed = _oge_row_to_parsed(row_data)
            if parsed:
                yield parsed
        return

    # Variant B: legacy dense or regional sparse xlsx export.
    current_layout = None
    current_column_map = None
    for row in sheet.iter_rows(min_row=1, values_only=True):
        if not row:
            continue
        if _is_oge_column_header_row(row):
            current_layout = _detect_oge_layout_from_header(row)
            current_column_map = _build_oge_column_map(row)
            continue
        layout = current_layout or _infer_oge_layout_from_data_row(row) or "legacy_dense"
        if current_layout is None:
            current_layout = layout
        column_map = current_column_map or _default_oge_column_map(layout)
        parsed = _parse_oge_row(row, column_map)
        if parsed:
            yield parsed


def _parse_oge_meta_from_xlsx_sheet(sheet):
    for row in sheet.iter_rows(min_row=1, max_row=25, values_only=True):
        row_values = [_safe_str(value) for value in row]
        header_text = _find_exam_header_in_row(row_values)
        if header_text:
            return parse_exam_header(header_text)
    return None


def _iter_oge_xlsx_exam_blocks(file_path):
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet = workbook.active
    current_exam_key = None
    current_layout = None
    current_column_map = None
    rows_by_exam = {}
    for row in sheet.iter_rows(min_row=1, values_only=True):
        header_text = _find_exam_header_in_row([_safe_str(value) for value in row] if row else [])
        if header_text:
            current_exam_key = parse_exam_header(header_text)
            rows_by_exam.setdefault(current_exam_key, [])
            current_layout = None
            current_column_map = None
            continue
        if not current_exam_key or not row:
            continue
        if _is_oge_column_header_row(row):
            current_layout = _detect_oge_layout_from_header(row)
            current_column_map = _build_oge_column_map(row)
            continue
        layout = current_layout or _infer_oge_layout_from_data_row(row) or "legacy_dense"
        if current_layout is None:
            current_layout = layout
        column_map = current_column_map or _default_oge_column_map(layout)
        parsed = _parse_oge_row(row, column_map)
        if parsed:
            rows_by_exam[current_exam_key].append(parsed)
    return rows_by_exam


def _iter_oge_xlsx_exam_blocks_stream(file_path):
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet = workbook.active
    current_exam_key = None
    current_layout = None
    current_column_map = None
    current_rows = []
    for row in sheet.iter_rows(min_row=1, values_only=True):
        header_text = _find_exam_header_in_row([_safe_str(value) for value in row] if row else [])
        if header_text:
            if current_exam_key and current_rows:
                yield current_exam_key, current_rows
            current_exam_key = parse_exam_header(header_text)
            current_rows = []
            current_layout = None
            current_column_map = None
            continue
        if not current_exam_key or not row:
            continue
        if _is_oge_column_header_row(row):
            current_layout = _detect_oge_layout_from_header(row)
            current_column_map = _build_oge_column_map(row)
            continue
        layout = current_layout or _infer_oge_layout_from_data_row(row) or "legacy_dense"
        if current_layout is None:
            current_layout = layout
        column_map = current_column_map or _default_oge_column_map(layout)
        parsed = _parse_oge_row(row, column_map)
        if parsed:
            current_rows.append(parsed)
    if current_exam_key and current_rows:
        yield current_exam_key, current_rows


def _iter_oge_xlsx_appeal_results_blocks_stream(file_path):
    """
    Compact appeal-results protocol:
    № | Код ОО | Класс | Код ППЭ | Аудитория | ФИО | Экзамен | апелляция | Первичный балл | Оценка
    Exam meta comes from the 'Экзамен' cell, e.g. Математика(2026.06.02).
    """
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet = workbook.active
    column_map = None
    rows_by_exam = {}
    for row in sheet.iter_rows(min_row=1, values_only=True):
        if not row:
            continue
        if _is_oge_appeal_results_header_row(row):
            column_map = dict(_default_oge_column_map("appeal_results"))
            built = _build_oge_column_map(row)
            for key in ("order", "school_code", "class_code", "station_code", "auditorium_code", "surname", "name", "patronymic"):
                if built.get(key) is not None:
                    column_map[key] = built[key]
            exam_idx = _find_header_index(list(row), "экзамен")
            if exam_idx is not None:
                column_map["exam_cell"] = exam_idx
            continue
        if column_map is None:
            continue
        # Refine score columns from a secondary header row under merged "Текущие".
        primary_idx = _find_header_index(list(row), "первичный балл")
        mark_idx = _find_header_index(list(row), "оценка")
        if primary_idx is not None or mark_idx is not None:
            if primary_idx is not None:
                column_map["primary_mark"] = primary_idx
            if mark_idx is not None:
                column_map["mark5"] = mark_idx
            continue

        parsed = _parse_oge_row(row, column_map)
        if not parsed:
            continue
        exam_raw = _safe_str(_oge_cell(row, column_map.get("exam_cell")))
        if not exam_raw:
            continue
        try:
            exam_key = parse_oge_exam_cell(exam_raw)
        except ValueError:
            continue
        rows_by_exam.setdefault(exam_key, []).append(parsed)

    for exam_key, parsed_rows in rows_by_exam.items():
        yield exam_key, parsed_rows


def _parse_oge_meta_from_xls_sheet(sheet):
    for row_idx in range(min(sheet.nrows, 25)):
        row_values = [_safe_str(sheet.cell_value(row_idx, col_idx)) for col_idx in range(min(sheet.ncols, 12))]
        header_text = _find_exam_header_in_row(row_values)
        if header_text:
            return parse_exam_header(header_text)
    return None


def _xls_row_values(sheet, row_idx: int) -> list:
    return [sheet.cell_value(row_idx, col_idx) for col_idx in range(sheet.ncols)]


def _iter_oge_xls_exam_blocks_stream(sheet):
    """Stream multi-exam regional/legacy OGE blocks from classic .xls sheets."""
    current_exam_key = None
    current_layout = None
    current_column_map = None
    current_rows = []
    for row_idx in range(sheet.nrows):
        row = _xls_row_values(sheet, row_idx)
        header_text = _find_exam_header_in_row([_safe_str(value) for value in row])
        if header_text:
            if current_exam_key and current_rows:
                yield current_exam_key, current_rows
            current_exam_key = parse_exam_header(header_text)
            current_rows = []
            current_layout = None
            current_column_map = None
            continue
        if not current_exam_key or not row:
            continue
        if _is_oge_column_header_row(row):
            current_layout = _detect_oge_layout_from_header(row)
            current_column_map = _build_oge_column_map(row)
            continue
        layout = current_layout or _infer_oge_layout_from_data_row(row) or "legacy_dense"
        if current_layout is None:
            current_layout = layout
        column_map = current_column_map or _default_oge_column_map(layout)
        parsed = _parse_oge_row(row, column_map)
        if parsed:
            current_rows.append(parsed)
    if current_exam_key and current_rows:
        yield current_exam_key, current_rows


def _iter_oge_xls_rows(sheet):
    if sheet.nrows < 2:
        return

    # Variant A: normalized tabular xls export with explicit English headers.
    first_row_headers = [_safe_str(sheet.cell_value(0, col_idx)) for col_idx in range(sheet.ncols)]
    if "SchoolCode" in first_row_headers and "PrimaryMark" in first_row_headers:
        headers = first_row_headers
        for row_idx in range(1, sheet.nrows):
            row_data = {
                headers[col_idx]: sheet.cell_value(row_idx, col_idx) if col_idx < sheet.ncols else ""
                for col_idx in range(len(headers))
                if headers[col_idx]
            }
            parsed = _oge_row_to_parsed(row_data)
            if parsed:
                yield parsed
        return

    # Variant B: legacy regional xls with merged header rows and sparse columns.
    for row_idx in range(0, sheet.nrows):
        # Data rows always begin with sequence number and school code.
        order_cell = sheet.cell_value(row_idx, 0) if sheet.ncols > 0 else ""
        school_cell = sheet.cell_value(row_idx, 1) if sheet.ncols > 1 else ""
        if not isinstance(order_cell, (int, float)) or not isinstance(school_cell, (int, float)):
            continue
        row_data = {
            "ColumnName": order_cell,
            "SchoolCode": school_cell,
            "ClassCode": sheet.cell_value(row_idx, 2) if sheet.ncols > 2 else "",
            "StationCode": sheet.cell_value(row_idx, 4) if sheet.ncols > 4 else "",
            "AuditoriumCode": sheet.cell_value(row_idx, 6) if sheet.ncols > 6 else "",
            "GovernmentCell": sheet.cell_value(row_idx, 8) if sheet.ncols > 8 else "",
            "SurnameCell": sheet.cell_value(row_idx, 10) if sheet.ncols > 10 else "",
            "NameCell": sheet.cell_value(row_idx, 12) if sheet.ncols > 12 else "",
            "SecondNameCell": sheet.cell_value(row_idx, 13) if sheet.ncols > 13 else "",
            "DocumentSeriesCode": sheet.cell_value(row_idx, 14) if sheet.ncols > 14 else "",
            "DocumentNumberCode": sheet.cell_value(row_idx, 16) if sheet.ncols > 16 else "",
            "TestResultBCell": sheet.cell_value(row_idx, 18) if sheet.ncols > 18 else "",
            "TestResultCCell": sheet.cell_value(row_idx, 21) if sheet.ncols > 21 else "",
            "PrimaryMark": sheet.cell_value(row_idx, 22) if sheet.ncols > 22 else "",
            "Mark5": sheet.cell_value(row_idx, 24) if sheet.ncols > 24 else "",
        }
        parsed = _oge_row_to_parsed(row_data)
        if parsed:
            yield parsed


def _parse_oge_exam_meta(file_path):
    basename = os.path.basename(file_path)
    match = OGE_FILE_PATTERN.search(basename)
    if not match:
        raise ValueError(f"Не удалось определить код и дату экзамена из имени файла: {basename}")
    code = match.group("code")
    exam_date = dt.datetime.strptime(match.group("date"), "%Y.%m.%d").date()
    subject = OGE_SUBJECT_CODE_MAP.get(code, f"ОГЭ код {code}")
    return code, subject, exam_date


def _build_school_map():
    school_map = {}
    for school in School.objects.only("id", "code"):
        school_map[school.code] = school.id
        normalized = _normalize_school_code(school.code)
        if normalized:
            school_map[normalized] = school.id
    return school_map


def _school_codes_filter(school_codes):
    if school_codes is None:
        return None
    return {_normalize_school_code(code) for code in school_codes if _normalize_school_code(code)}


def _upsert_exam_data(
    exam: Exam,
    parsed_rows: Iterable[ParsedRow],
    school_map=None,
    school_codes=None,
    stats: ParseStats | None = None,
):
    school_map = school_map or _build_school_map()
    allowed_codes = _school_codes_filter(school_codes)
    normalized_rows = []
    unique_student_keys = set()
    first_parsed_by_key = {}
    school_ids = set()
    external_ids = set()

    for parsed in parsed_rows:
        row_code = _normalize_school_code(parsed.school_code)
        if allowed_codes is not None and row_code not in allowed_codes:
            if stats is not None:
                stats.skipped_other_school += 1
            continue
        school_id = school_map.get(row_code) or school_map.get(parsed.school_code)
        if not school_id:
            if stats is not None:
                stats.skipped_unknown_school += 1
            continue
        key = (school_id, parsed.student_id)
        unique_student_keys.add(key)
        first_parsed_by_key.setdefault(key, parsed)
        school_ids.add(school_id)
        external_ids.add(parsed.student_id)
        normalized_rows.append((school_id, parsed))

    existing_students = Student.objects.filter(
        school_id__in=school_ids,
        external_id__in=external_ids,
    ).only("id", "school_id", "external_id")
    student_map = {(s.school_id, s.external_id): s.id for s in existing_students}

    to_create_students = []
    for school_id, student_id in unique_student_keys:
        if (school_id, student_id) not in student_map:
            parsed = first_parsed_by_key[(school_id, student_id)]
            to_create_students.append(
                Student(
                    school_id=school_id,
                    external_id=student_id,
                    full_name=parsed.full_name,
                    grade=parsed.grade,
                )
            )
    Student.objects.bulk_create(to_create_students, batch_size=2000)

    refreshed_students = Student.objects.filter(
        school_id__in=school_ids,
        external_id__in=external_ids,
    ).only("id", "school_id", "external_id")
    for student in refreshed_students:
        student_map[(student.school_id, student.external_id)] = student.id

    student_updates = {}
    for school_id, parsed in normalized_rows:
        student_id = student_map[(school_id, parsed.student_id)]
        student_updates[student_id] = (parsed.full_name, parsed.grade)
    update_batch = [
        Student(id=s_id, full_name=full_name, grade=grade)
        for s_id, (full_name, grade) in student_updates.items()
    ]
    Student.objects.bulk_update(update_batch, ["full_name", "grade"], batch_size=2000)

    # one result per student per exam; if duplicate rows exist, keep max score row
    best_rows = {}
    for school_id, parsed in normalized_rows:
        key = (school_id, parsed.student_id)
        existing = best_rows.get(key)
        if existing is None or parsed.total_score > existing.total_score:
            best_rows[key] = parsed

    if allowed_codes is not None:
        # Replace only schools present in this file (and allowed), so a district
        # upload does not wipe results for schools absent from the protocol.
        target_school_ids = school_ids
        if target_school_ids:
            ExamResult.objects.filter(exam=exam, student__school_id__in=target_school_ids).delete()
            TaskResult.objects.filter(exam=exam, student__school_id__in=target_school_ids).delete()
    else:
        ExamResult.objects.filter(exam=exam).delete()
        TaskResult.objects.filter(exam=exam).delete()

    exam_results = []
    task_results = []
    for (school_id, ext_id), parsed in best_rows.items():
        student_id = student_map[(school_id, ext_id)]
        exam_results.append(
            ExamResult(
                student_id=student_id,
                exam=exam,
                school_code=parsed.school_code,
                student_name=parsed.full_name,
                short_answer_tasks=parsed.short_answer_tasks,
                long_answer_tasks=parsed.long_answer_tasks,
                primary_score=parsed.primary_score,
                score=parsed.score,
                total_score=parsed.total_score,
                passed=(parsed.score >= 3) if exam.exam_type == "oge" else (parsed.total_score > 0),
                short_answer_raw=parsed.short_answer_tasks,
                source_row=parsed.source_row,
            )
        )
        short_mask = parsed.short_answer_tasks or ""
        for idx, token in enumerate(short_mask, start=1):
            task_results.append(
                TaskResult(
                    student_id=student_id,
                    exam=exam,
                    task_number=idx,
                    value=token,
                )
            )
        part2_start = len(short_mask) + 1 if short_mask else 1
        for task_number, token in parse_long_answer_mask(parsed.long_answer_tasks, part2_start):
            task_results.append(
                TaskResult(
                    student_id=student_id,
                    exam=exam,
                    task_number=task_number,
                    value=token,
                )
            )

    ExamResult.objects.bulk_create(exam_results, batch_size=2000)
    for i in range(0, len(task_results), 10000):
        TaskResult.objects.bulk_create(task_results[i : i + 10000], batch_size=10000)
    if stats is not None:
        stats.results_imported += len(exam_results)


def parse_ege(file_path, school_codes=None):
    stats = ParseStats()
    school_map = _build_school_map()
    rows_by_exam = _iter_ege_exam_blocks(file_path, start_row=1)
    exam_ids = []
    for (code, subject, exam_date), parsed_rows in rows_by_exam.items():
        exam, _ = Exam.objects.get_or_create(
            exam_type="ege",
            code=code,
            exam_date=exam_date,
            defaults={"subject": subject, "year": exam_date.year},
        )
        _upsert_exam_data(
            exam,
            parsed_rows,
            school_map=school_map,
            school_codes=school_codes,
            stats=stats,
        )
        exam_ids.append(exam.id)
        stats.exams_processed += 1
    if school_codes is not None and stats.results_imported == 0:
        raise ValueError(
            "В файле не найдены данные по кодам ОО вашей организации или района. "
            "Проверьте код ОО в протоколе."
        )
    return exam_ids, stats


def _process_oge_exam_block(code, subject, exam_date, parsed_rows, school_map, school_codes, stats):
    exam, _ = Exam.objects.get_or_create(
        exam_type="oge",
        code=code,
        exam_date=exam_date,
        defaults={"subject": subject, "year": exam_date.year},
    )
    _upsert_exam_data(
        exam,
        parsed_rows,
        school_map=school_map,
        school_codes=school_codes,
        stats=stats,
    )
    stats.exams_processed += 1
    return exam.id


def parse_oge(file_path, school_codes=None):
    stats = ParseStats()
    school_map = _build_school_map()
    exam_ids = []
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        code, subject, exam_date = _parse_oge_exam_meta(file_path)
        parsed_rows = _iter_oge_csv_rows(file_path)
        exam_ids.append(
            _process_oge_exam_block(code, subject, exam_date, parsed_rows, school_map, school_codes, stats)
        )
    elif ext == ".xls":
        workbook = xlrd.open_workbook(file_path)
        sheet = workbook.sheet_by_index(0)
        for (code, subject, exam_date), parsed_rows in _iter_oge_xls_exam_blocks_stream(sheet):
            exam_ids.append(
                _process_oge_exam_block(code, subject, exam_date, parsed_rows, school_map, school_codes, stats)
            )
        if not exam_ids:
            # Fallback: single-exam / legacy column layouts without exam title rows.
            parsed_meta = _parse_oge_meta_from_xls_sheet(sheet)
            if parsed_meta:
                code, subject, exam_date = parsed_meta
            else:
                code, subject, exam_date = _parse_oge_exam_meta(file_path)
            parsed_rows = _iter_oge_xls_rows(sheet)
            exam_ids.append(
                _process_oge_exam_block(code, subject, exam_date, parsed_rows, school_map, school_codes, stats)
            )
    elif ext == ".xlsx":
        if _is_oge_appeal_results_xlsx(file_path):
            for (code, subject, exam_date), parsed_rows in _iter_oge_xlsx_appeal_results_blocks_stream(file_path):
                exam_ids.append(
                    _process_oge_exam_block(code, subject, exam_date, parsed_rows, school_map, school_codes, stats)
                )
        else:
            for (code, subject, exam_date), parsed_rows in _iter_oge_xlsx_exam_blocks_stream(file_path):
                exam_ids.append(
                    _process_oge_exam_block(code, subject, exam_date, parsed_rows, school_map, school_codes, stats)
                )
            if not exam_ids:
                workbook = load_workbook(file_path, read_only=True, data_only=True)
                sheet = workbook.active
                parsed_meta = _parse_oge_meta_from_xlsx_sheet(sheet)
                if parsed_meta:
                    code, subject, exam_date = parsed_meta
                else:
                    code, subject, exam_date = _parse_oge_exam_meta(file_path)
                parsed_rows = _iter_oge_xlsx_rows(file_path)
                exam_ids.append(
                    _process_oge_exam_block(code, subject, exam_date, parsed_rows, school_map, school_codes, stats)
                )
    else:
        raise ValueError(f"Неподдерживаемый формат файла ОГЭ: {ext}")

    if school_codes is not None and stats.results_imported == 0:
        raise ValueError(
            "В файле не найдены данные по кодам ОО вашей организации или района. "
            "Проверьте код ОО в протоколе."
        )
    return exam_ids, stats
