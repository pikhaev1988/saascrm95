import io

from django.test import TestCase
from openpyxl import Workbook

from uploads.parsers import (
    _build_ege_column_map,
    _build_oge_column_map,
    _detect_ege_layout_from_header,
    _iter_ege_exam_blocks,
    _iter_oge_xlsx_appeal_results_blocks_stream,
    _iter_oge_xlsx_exam_blocks_stream,
    parse_exam_header,
    parse_oge_exam_cell,
)
from uploads.sample_protocols import build_ege_sample_xlsx, build_oge_sample_xlsx


def _build_regional_ege_sample_xlsx(school_code: str = "13450", msu_code: str = "213") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Page 1"
    ws.append(["Протокол проверки"])
    ws.append(["20 - Татарстан Республика"])
    ws.append([None])
    ws.append([None])
    ws.append([None])
    ws.append(["04 - Химия 2026.06.01"])
    ws.append([None])
    ws.append(
        [
            "№",
            None,
            "Код МСУ",
            None,
            "Код ОО",
            None,
            "Класс",
            None,
            "Код ППЭ",
            None,
            "Аудитория",
            None,
            "Фамилия",
            "Имя",
            "Отчество",
            None,
            "Задания с кратким ответом",
            None,
            "Задания с развёрнутым ответом",
            None,
            "Первичный балл",
            None,
            "Тестовый балл",
            None,
        ]
    )
    ws.append(
        [
            1,
            None,
            int(msu_code),
            None,
            int(school_code),
            None,
            "11",
            None,
            1301,
            None,
            108,
            None,
            "Иванов",
            "Иван",
            "Иванович",
            None,
            "++---+-------------",
            None,
            "1(1)2(3)1(2)",
            None,
            15,
            None,
            35,
            None,
        ]
    )
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


class EgeProtocolParserTests(TestCase):
    def test_parse_compact_sample_protocol(self):
        content = build_ege_sample_xlsx("12345")
        blocks = _iter_ege_exam_blocks(io.BytesIO(content))
        self.assertEqual(len(blocks), 1)
        (_, subject, exam_date), rows = next(iter(blocks.items()))
        self.assertEqual(parse_exam_header("01 - Русский язык 2025.05.30")[1], subject)
        self.assertEqual(len(rows), 2)
        first = rows[0]
        self.assertEqual(first.school_code, "12345")
        self.assertEqual(first.student_id, "000001")
        self.assertEqual(first.primary_score, 15.0)
        self.assertEqual(first.total_score, 35.0)
        self.assertIn("++", first.short_answer_tasks)

    def test_build_regional_column_map_uses_document_headers(self):
        header_row = (
            "№",
            None,
            "Код МСУ",
            None,
            "Код ОО",
            None,
            "Класс",
            None,
            "Код ППЭ",
            None,
            "Аудитория",
            None,
            "Фамилия",
            "Имя",
            "Отчество",
            None,
            "Задания с кратким ответом",
            None,
            "Задания с развёрнутым ответом",
            None,
            "Первичный балл",
            None,
            "Тестовый балл",
            None,
        )
        column_map = _build_ege_column_map(header_row)
        self.assertEqual(column_map["school_code"], 4)
        self.assertIsNone(column_map["participant_code"])
        self.assertEqual(column_map["surname"], 12)
        self.assertEqual(column_map["total_score"], 22)

    def test_parse_regional_protocol(self):
        content = _build_regional_ege_sample_xlsx("13450")
        blocks = _iter_ege_exam_blocks(io.BytesIO(content))
        self.assertEqual(len(blocks), 1)
        (code, subject, _), rows = next(iter(blocks.items()))
        self.assertEqual(code, "04")
        self.assertEqual(subject, "Химия")
        self.assertEqual(len(rows), 1)
        first = rows[0]
        self.assertEqual(first.school_code, "13450")
        self.assertEqual(first.student_id, "Иванов Иван Иванович")
        self.assertEqual(first.full_name, "Иванов Иван Иванович")
        self.assertEqual(first.primary_score, 15.0)
        self.assertEqual(first.total_score, 35.0)


def _build_regional_oge_sample_xlsx(school_code: str = "13450", msu_code: str = "213") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Page 1"
    ws.append(["02 - Математика 2026.06.02"])
    ws.append(
        [
            "№",
            "Код ОО",
            "Класс",
            None,
            "Код ППЭ",
            None,
            "Аудитория",
            None,
            "Код МСУ",
            None,
            "Фамилия",
            None,
            "Имя",
            "Отчество",
            "Задания с кратким ответом",
            None,
            None,
            "Задания с развёрнутым ответом",
            "Первичный балл",
            None,
            "Оценка",
        ]
    )
    ws.append(
        [
            1,
            int(school_code),
            "9а",
            None,
            1316,
            None,
            203,
            None,
            int(msu_code),
            None,
            "Иванов",
            None,
            "Иван",
            "Иванович",
            "++-+++++-++++++++++",
            None,
            None,
            "0(2)0(2)0(2)0(2)0(2)0(2)",
            17,
            None,
            4,
        ]
    )
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _build_oge_appeal_results_sample_xlsx(school_code: str = "18579") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Page 1"
    ws.append(["Протокол проверки результатов  государственной итоговой аттестации 2026 г."])
    ws.append(["20 - Чеченская Республика"])
    ws.append([None])
    ws.append(
        [
            "№",
            "Код ОО",
            "Класс",
            "Код ППЭ",
            "Аудитория",
            "Фамилия",
            "Имя",
            "Отчество",
            "Экзамен",
            "Состояние апелляции",
            "Текущие",
            None,
        ]
    )
    ws.append(
        [
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            "Первичный балл",
            "Оценка",
        ]
    )
    ws.append(
        [
            1,
            school_code,
            "А",
            "503",
            "0204",
            "Абдурзакова",
            "Марха",
            "Магомедовна",
            "Математика(2026.06.02)",
            "Апелляция удовлетворена",
            "22",
            "5",
        ]
    )
    ws.append(
        [
            2,
            "18560",
            "9А",
            "503",
            "0207",
            "Сатуев",
            "Сайфула",
            "Алиханович",
            "Математика(2026.06.02)",
            "Апелляция удовлетворена",
            "12",
            "3",
        ]
    )
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


class OgeProtocolParserTests(TestCase):
    def test_parse_legacy_oge_sample_protocol(self):
        content = build_oge_sample_xlsx("12345")
        blocks = list(_iter_oge_xlsx_exam_blocks_stream(io.BytesIO(content)))
        self.assertEqual(len(blocks), 1)
        (_, subject, exam_date), rows = blocks[0]
        self.assertEqual(subject, "Русский язык")
        self.assertEqual(len(rows), 2)
        first = rows[0]
        self.assertEqual(first.school_code, "12345")

    def test_build_regional_oge_column_map_uses_document_headers(self):
        header_row = (
            "№",
            "Код ОО",
            "Класс",
            None,
            "Код ППЭ",
            None,
            "Аудитория",
            None,
            "Код МСУ",
            None,
            "Фамилия",
            None,
            "Имя",
            "Отчество",
            "Задания с кратким ответом",
            None,
            None,
            "Задания с развёрнутым ответом",
            "Первичный балл",
            None,
            "Оценка",
        )
        column_map = _build_oge_column_map(header_row)
        self.assertEqual(column_map["school_code"], 1)
        self.assertEqual(column_map["msu_code"], 8)
        self.assertEqual(column_map["surname"], 10)
        self.assertEqual(column_map["mark5"], 20)

    def test_parse_regional_oge_protocol(self):
        content = _build_regional_oge_sample_xlsx("13450")
        blocks = list(_iter_oge_xlsx_exam_blocks_stream(io.BytesIO(content)))
        self.assertEqual(len(blocks), 1)
        (code, subject, _), rows = blocks[0]
        self.assertEqual(code, "02")
        self.assertEqual(subject, "Математика")
        self.assertEqual(len(rows), 1)
        first = rows[0]
        self.assertEqual(first.school_code, "13450")
        self.assertEqual(first.student_id, "Иванов Иван Иванович")
        self.assertEqual(first.full_name, "Иванов Иван Иванович")
        self.assertEqual(first.primary_score, 17.0)
        self.assertEqual(first.score, 4.0)
        self.assertIn("++", first.short_answer_tasks)
        self.assertEqual(first.long_answer_tasks, "0(2)0(2)0(2)0(2)0(2)0(2)")

    def test_parse_oge_exam_cell(self):
        code, subject, exam_date = parse_oge_exam_cell("Математика(2026.06.02)")
        self.assertEqual(code, "02")
        self.assertEqual(subject, "Математика")
        self.assertEqual(str(exam_date), "2026-06-02")

    def test_parse_oge_appeal_results_protocol(self):
        content = _build_oge_appeal_results_sample_xlsx("18579")
        blocks = list(_iter_oge_xlsx_appeal_results_blocks_stream(io.BytesIO(content)))
        self.assertEqual(len(blocks), 1)
        (code, subject, exam_date), rows = blocks[0]
        self.assertEqual(code, "02")
        self.assertEqual(subject, "Математика")
        self.assertEqual(str(exam_date), "2026-06-02")
        self.assertEqual(len(rows), 2)
        first = rows[0]
        self.assertEqual(first.school_code, "18579")
        self.assertEqual(first.full_name, "Абдурзакова Марха Магомедовна")
        self.assertEqual(first.primary_score, 22.0)
        self.assertEqual(first.score, 5.0)
        self.assertEqual(first.short_answer_tasks, "")
        self.assertEqual(rows[1].score, 3.0)

    def test_parse_long_answer_mask_from_protocol(self):
        from users.task_topics import parse_long_answer_mask

        rows = parse_long_answer_mask("0(2)0(2)0(2)0(2)0(2)0(2)", 20)
        self.assertEqual(len(rows), 6)
        self.assertEqual(rows[0], (20, "0"))
        self.assertEqual(rows[5], (25, "0"))


class SchoolUploadRevertTests(TestCase):
    def setUp(self):
        from organizations.models import District, Ministry, School
        from users.models import User

        self.ministry = Ministry.objects.create(name="Тестовое министерство")
        self.district = District.objects.create(ministry=self.ministry, code="01", name="Тестовый район")
        self.school = School.objects.create(district=self.district, code="12345", name="Тестовая школа")
        self.user = User.objects.create_user(
            username="school_user",
            password="pass",
            role="school",
            school=self.school,
        )

    def test_revert_school_upload_removes_school_results(self):
        import tempfile

        from exams.models import Exam, ExamResult
        from uploads.models import UploadSession
        from uploads.parsers import parse_ege
        from uploads.services import link_upload_exams, revert_school_upload

        content = build_ege_sample_xlsx(self.school.code)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(content)
            tmp_path = tmp.name

        exam_ids, stats = parse_ege(tmp_path, school_codes=[self.school.code])
        self.assertGreater(stats.results_imported, 0)

        session = UploadSession.objects.create(
            uploaded_by=self.user,
            school=self.school,
            exam_type="ege",
            status="done",
            results_imported=stats.results_imported,
            exams_processed=stats.exams_processed,
        )
        link_upload_exams(session, exam_ids)

        exam = Exam.objects.get(pk=exam_ids[0])
        self.assertTrue(
            ExamResult.objects.filter(exam=exam, student__school=self.school).exists()
        )

        result = revert_school_upload(session)
        self.assertEqual(result["exams_affected"], 1)
        self.assertFalse(
            ExamResult.objects.filter(exam=exam, student__school=self.school).exists()
        )
        session.refresh_from_db()
        self.assertIsNotNone(session.reverted_at)

    def test_delete_school_exam_results(self):
        import tempfile

        from exams.models import Exam, ExamResult
        from uploads.parsers import parse_ege
        from uploads.services import delete_school_exam_results

        content = build_ege_sample_xlsx(self.school.code)
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(content)
            tmp_path = tmp.name

        exam_ids, stats = parse_ege(tmp_path, school_codes=[self.school.code])
        exam = Exam.objects.get(pk=exam_ids[0])
        self.assertTrue(ExamResult.objects.filter(exam=exam, student__school=self.school).exists())

        result = delete_school_exam_results(self.school, exam)
        self.assertGreater(result["results_removed"], 0)
        self.assertFalse(ExamResult.objects.filter(exam=exam, student__school=self.school).exists())

    def test_revert_twice_raises_error(self):
        from uploads.models import UploadSession
        from uploads.services import revert_school_upload

        session = UploadSession.objects.create(
            uploaded_by=self.user,
            school=self.school,
            exam_type="ege",
            status="done",
            results_imported=1,
            exams_processed=1,
        )
        revert_school_upload(session)
        with self.assertRaisesMessage(ValueError, "уже отменена"):
            revert_school_upload(session)

    def test_revert_view_requires_owner(self):
        from django.test import Client
        from uploads.models import UploadSession
        from users.models import User

        other_user = User.objects.create_user(
            username="other_school",
            password="pass",
            role="school",
            school=self.school,
        )
        session = UploadSession.objects.create(
            uploaded_by=self.user,
            school=self.school,
            exam_type="ege",
            status="done",
        )
        client = Client()
        client.force_login(other_user)
        response = client.post(f"/cabinet/upload/ege/revert/{session.id}/")
        self.assertEqual(response.status_code, 404)


class DistrictUploadFilterTests(TestCase):
    def setUp(self):
        from organizations.models import District, Ministry, School
        from users.models import User

        self.ministry = Ministry.objects.create(name="Мин")
        self.district = District.objects.create(ministry=self.ministry, code="D1", name="Район 1")
        self.other_district = District.objects.create(ministry=self.ministry, code="D2", name="Район 2")
        self.school_a = School.objects.create(district=self.district, code="11111", name="Школа А")
        self.school_b = School.objects.create(district=self.district, code="22222", name="Школа Б")
        self.school_other = School.objects.create(
            district=self.other_district, code="99999", name="Чужая школа"
        )
        self.user = User.objects.create_user(
            username="district_user",
            password="pass",
            role="district",
            district=self.district,
        )

    def _write_multi_school_ege(self, codes):
        import tempfile

        wb = Workbook()
        ws = wb.active
        ws.append(["01 - Русский язык 2025.05.30"])
        ws.append(
            [
                "№",
                "Код ППЭ",
                "Код ОО",
                "Класс",
                "Код ООП",
                "Профиль",
                "Фамилия",
                "Имя",
                "Отчество",
                "Вариант",
                "Код участника",
                "№ бланка",
                "Ответы в краткой форме",
                "Ответы в развернутой форме",
                "Первичный балл",
                "Итоговый балл",
            ]
        )
        for idx, code in enumerate(codes, start=1):
            ws.append(
                [
                    idx,
                    100,
                    int(code) if str(code).isdigit() else code,
                    "11",
                    101,
                    102,
                    f"Фам{idx}",
                    f"Имя{idx}",
                    "Отч",
                    1,
                    f"{idx:06d}",
                    f"{idx:06d}",
                    "++---+-------------",
                    "1(1)2(3)1(2)",
                    15,
                    35,
                ]
            )
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb.save(tmp.name)
            return tmp.name

    def test_district_codes_skip_other_district_schools(self):
        import tempfile

        from exams.models import ExamResult
        from uploads.parsers import parse_ege

        path = self._write_multi_school_ege(
            [self.school_a.code, self.school_other.code, self.school_b.code]
        )
        district_codes = [self.school_a.code, self.school_b.code]
        exam_ids, stats = parse_ege(path, school_codes=district_codes)

        self.assertEqual(stats.results_imported, 2)
        self.assertEqual(stats.skipped_other_school, 1)
        self.assertTrue(
            ExamResult.objects.filter(exam_id=exam_ids[0], student__school=self.school_a).exists()
        )
        self.assertTrue(
            ExamResult.objects.filter(exam_id=exam_ids[0], student__school=self.school_b).exists()
        )
        self.assertFalse(
            ExamResult.objects.filter(exam_id=exam_ids[0], student__school=self.school_other).exists()
        )

    def test_district_upload_does_not_wipe_absent_school(self):
        from exams.models import ExamResult
        from uploads.parsers import parse_ege

        path_b = self._write_multi_school_ege([self.school_b.code])
        parse_ege(path_b, school_codes=[self.school_b.code])
        self.assertTrue(ExamResult.objects.filter(student__school=self.school_b).exists())

        path_a = self._write_multi_school_ege([self.school_a.code])
        district_codes = [self.school_a.code, self.school_b.code]
        parse_ege(path_a, school_codes=district_codes)

        self.assertTrue(ExamResult.objects.filter(student__school=self.school_a).exists())
        self.assertTrue(
            ExamResult.objects.filter(student__school=self.school_b).exists(),
            "Школа без строк в файле не должна терять результаты при районной загрузке",
        )

    def test_revert_district_upload(self):
        from exams.models import Exam, ExamResult
        from uploads.models import UploadSession
        from uploads.parsers import parse_ege
        from uploads.services import link_upload_exams, revert_district_upload

        path = self._write_multi_school_ege([self.school_a.code, self.school_b.code])
        exam_ids, stats = parse_ege(path, school_codes=[self.school_a.code, self.school_b.code])
        session = UploadSession.objects.create(
            uploaded_by=self.user,
            district=self.district,
            exam_type="ege",
            status="done",
            results_imported=stats.results_imported,
            exams_processed=stats.exams_processed,
        )
        link_upload_exams(session, exam_ids)
        exam = Exam.objects.get(pk=exam_ids[0])
        self.assertEqual(
            ExamResult.objects.filter(exam=exam, student__school__district=self.district).count(),
            2,
        )

        result = revert_district_upload(session)
        self.assertEqual(result["exams_affected"], 1)
        self.assertEqual(
            ExamResult.objects.filter(exam=exam, student__school__district=self.district).count(),
            0,
        )
        session.refresh_from_db()
        self.assertIsNotNone(session.reverted_at)

    def test_district_upload_page_requires_district_role(self):
        from django.test import Client
        from users.models import User

        school_user = User.objects.create_user(
            username="school_only",
            password="pass",
            role="school",
            school=self.school_a,
        )
        client = Client()
        client.force_login(school_user)
        response = client.get("/cabinet/district/upload/ege/")
        self.assertEqual(response.status_code, 302)

        client.force_login(self.user)
        response = client.get("/cabinet/district/upload/ege/")
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, self.district.name)

