from openpyxl import load_workbook
from django.db.models import Count, Q
from django.utils import timezone

from exams.models import Exam, ExamResult, TaskResult
from organizations.models import District, School
from uploads.models import UploadSession


def link_upload_exams(session: UploadSession, exam_ids) -> None:
    ids = [exam_id for exam_id in exam_ids if exam_id]
    if ids:
        session.exams.add(*ids)


def delete_school_exam_results(school: School, exam: Exam) -> dict:
    school_filter = {"student__school_id": school.id}
    results_qs = ExamResult.objects.filter(exam=exam, **school_filter)
    results_removed = results_qs.count()
    if results_removed == 0:
        raise ValueError("У вашей школы нет результатов по этому экзамену.")
    results_qs.delete()
    TaskResult.objects.filter(exam=exam, **school_filter).delete()
    return {"results_removed": results_removed, "exam_id": exam.id}


def get_school_exams_with_results(school: School, exam_type: str):
    return (
        Exam.objects.filter(exam_type=exam_type, results__student__school=school)
        .annotate(
            school_results_count=Count(
                "results",
                filter=Q(results__student__school_id=school.id),
            )
        )
        .order_by("-exam_date", "subject")
        .distinct()
    )


def revert_school_upload(session: UploadSession) -> dict:
    if session.reverted_at:
        raise ValueError("Эта загрузка уже отменена.")
    if session.status != "done":
        raise ValueError("Можно отменить только успешную загрузку.")
    if not session.school_id:
        raise ValueError("Отмена доступна только для загрузок из кабинета школы.")

    exam_ids = list(session.exams.values_list("id", flat=True))
    results_removed = 0
    for exam in Exam.objects.filter(id__in=exam_ids):
        deleted = delete_school_exam_results(session.school, exam)
        results_removed += deleted["results_removed"]

    session.reverted_at = timezone.now()
    session.save(update_fields=["reverted_at"])
    return {"exams_affected": len(exam_ids), "results_removed": results_removed}


def delete_district_exam_results(district: District, exam: Exam) -> dict:
    district_filter = {"student__school__district_id": district.id}
    results_qs = ExamResult.objects.filter(exam=exam, **district_filter)
    results_removed = results_qs.count()
    if results_removed == 0:
        raise ValueError("У школ вашего района нет результатов по этому экзамену.")
    results_qs.delete()
    TaskResult.objects.filter(exam=exam, **district_filter).delete()
    return {"results_removed": results_removed, "exam_id": exam.id}


def get_district_exams_with_results(district: District, exam_type: str):
    return (
        Exam.objects.filter(exam_type=exam_type, results__student__school__district=district)
        .annotate(
            school_results_count=Count(
                "results",
                filter=Q(results__student__school__district_id=district.id),
            )
        )
        .order_by("-exam_date", "subject")
        .distinct()
    )


def revert_district_upload(session: UploadSession) -> dict:
    if session.reverted_at:
        raise ValueError("Эта загрузка уже отменена.")
    if session.status != "done":
        raise ValueError("Можно отменить только успешную загрузку.")
    if not session.district_id:
        raise ValueError("Отмена доступна только для загрузок из кабинета района.")

    exam_ids = list(session.exams.values_list("id", flat=True))
    results_removed = 0
    exams_affected = 0
    for exam in Exam.objects.filter(id__in=exam_ids):
        try:
            deleted = delete_district_exam_results(session.district, exam)
        except ValueError:
            continue
        results_removed += deleted["results_removed"]
        exams_affected += 1

    session.reverted_at = timezone.now()
    session.save(update_fields=["reverted_at"])
    return {"exams_affected": exams_affected, "results_removed": results_removed}


def _normalize_header(value):
    return str(value or "").strip().lower()


def _choose_columns(headers):
    normalized = [_normalize_header(h) for h in headers]
    school_code_idx = next((i for i, h in enumerate(normalized) if "код" in h and "оо" in h), 0)
    school_name_idx = next((i for i, h in enumerate(normalized) if "наимен" in h), 1)
    district_code_idx = next(
        (i for i, h in enumerate(normalized) if ("мсу" in h or "район" in h) and "код" in h),
        2,
    )
    district_name_idx = next(
        (i for i, h in enumerate(normalized) if ("мсу" in h or "район" in h) and "наимен" in h),
        None,
    )
    return school_code_idx, school_name_idx, district_code_idx, district_name_idx


def import_organizations_from_excel(file_path, ministry):
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(min_row=1, values_only=True)
    header = next(rows, None)
    if not header:
        return {"districts_created": 0, "schools_created": 0, "schools_updated": 0}

    school_code_idx, school_name_idx, district_code_idx, district_name_idx = _choose_columns(header)
    districts_created = 0
    schools_created = 0
    schools_updated = 0
    for row in rows:
        if not row:
            continue
        school_code = str(row[school_code_idx] if len(row) > school_code_idx else "").strip()
        school_name = str(row[school_name_idx] if len(row) > school_name_idx else "").strip()
        district_code = str(row[district_code_idx] if len(row) > district_code_idx else "").strip()
        district_name = (
            str(row[district_name_idx]).strip()
            if district_name_idx is not None and len(row) > district_name_idx and row[district_name_idx]
            else f"Район {district_code}"
        )
        if not school_code or not district_code:
            continue

        district, district_created = District.objects.get_or_create(
            ministry=ministry,
            code=district_code,
            defaults={"name": district_name},
        )
        if district_created:
            districts_created += 1
        if district.name != district_name:
            district.name = district_name
            district.save(update_fields=["name"])
        school, created = School.objects.update_or_create(
            code=school_code,
            defaults={"district": district, "name": school_name},
        )
        if created:
            schools_created += 1
        else:
            schools_updated += 1

    return {
        "districts_created": districts_created,
        "schools_created": schools_created,
        "schools_updated": schools_updated,
    }
