"""
Синхронизация тем ЕГЭ/ОГЭ по русскому языку из Excel в data/ege_2026_enriched.json.
Запуск: PYTHONUTF8=1 python scripts/sync_russian_topics_from_xlsx.py
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
ENRICHED = ROOT / "data" / "ege_2026_enriched.json"
XLSX = Path.home() / "Desktop" / "Новая папка (5)" / "Русский язык.xlsx"


def norm(s: str | None) -> str:
    if not s:
        return ""
    t = str(s).replace("\r\n", "\n").replace("\r", "\n")
    t = re.sub(r"\s+", " ", t).strip()
    # переносы в ячейках Excel часто дают «об- щее» → «обще»
    t = re.sub(r"(\w)-\s+(\w)", r"\1\2", t)
    return t


def parse_grades_from_program_text(text: str) -> list[int]:
    """Извлекает номера классов из текста вида «5 класс», «5–9 классы»."""
    if not text:
        return []
    grades: set[int] = set()
    for m in re.finditer(r"(\d+)\s*[–-]\s*(\d+)\s*класс", text, flags=re.I):
        a, b = int(m.group(1)), int(m.group(2))
        lo, hi = min(a, b), max(a, b)
        for g in range(lo, hi + 1):
            if 1 <= g <= 11:
                grades.add(g)
    for m in re.finditer(r"(\d+)\s*класс", text, flags=re.I):
        g = int(m.group(1))
        if 1 <= g <= 11:
            grades.add(g)
    return sorted(grades)


def read_xlsx_rows() -> dict[int, tuple[str, str]]:
    from openpyxl import load_workbook

    if not XLSX.exists():
        raise SystemExit(f"Файл не найден: {XLSX}")
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    out: dict[int, tuple[str, str]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        try:
            num = int(float(row[0]))
        except (TypeError, ValueError):
            continue
        col_ege = norm(row[1]) if len(row) > 1 else ""
        col_oge = norm(row[2]) if len(row) > 2 else ""
        out[num] = (col_ege, col_oge)
    return out


def main() -> None:
    xlsx_map = read_xlsx_rows()
    raw = ENRICHED.read_text(encoding="utf-8")
    data = json.loads(raw)
    subjects = data.get("subjects", [])
    rus = next((s for s in subjects if s.get("subject") == "russian"), None)
    if not rus:
        raise SystemExit("В JSON нет subject=russian")

    old_by_task = {int(t["task"]): t for t in rus.get("tasks", []) if t.get("task") is not None}
    new_tasks: list[dict] = []

    for task_num in sorted(set(old_by_task) | set(xlsx_map)):
        old = old_by_task.get(task_num, {})
        ege_txt, oge_txt = xlsx_map.get(task_num, ("", ""))
        topic_ege = ege_txt or (old.get("topic") or "")
        topic_oge = oge_txt or topic_ege

        grades_ege = parse_grades_from_program_text(ege_txt) if ege_txt else []
        grades_oge = parse_grades_from_program_text(oge_txt) if oge_txt else []
        if not grades_ege:
            grades_ege = list(old.get("grade_range") or [])
        if not grades_oge:
            grades_oge = list(old.get("grade_range") or [5, 6, 7, 8, 9])

        entry: dict = {
            "task": task_num,
            "topic": topic_ege,
            "topic_oge": topic_oge,
            "grade_range": grades_ege,
            "grade_range_oge": grades_oge,
            "source": "school_program_xlsx",
        }
        if old.get("skill"):
            entry["skill"] = old["skill"]
        new_tasks.append(entry)

    rus["tasks"] = new_tasks
    ENRICHED.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"OK: обновлено {len(new_tasks)} заданий в {ENRICHED}")


if __name__ == "__main__":
    main()
