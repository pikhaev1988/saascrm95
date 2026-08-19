from collections import defaultdict

from django.db import transaction
from openpyxl import load_workbook

from exams.models import Exam, ExamResult, Student, TaskResult
from organizations.models import School
from uploads.parsers import parse_exam_header


FILE_PATH = r"g:\Результаты ЕГЭ\2023 год ЕГЭ.xlsx"


def run():
    workbook = load_workbook(FILE_PATH, read_only=True, data_only=True)
    sheet = workbook.active
    code, subject, exam_date = parse_exam_header(str(sheet["A1"].value))
    exam, _ = Exam.objects.get_or_create(
        exam_type="ege",
        code=code,
        exam_date=exam_date,
        defaults={"subject": subject, "year": exam_date.year},
    )

    school_by_code = {str(s.code): s.id for s in School.objects.only("id", "code")}
    parsed_rows = []
    school_ids = set()
    ext_ids = set()
    for row in sheet.iter_rows(min_row=3, values_only=True):
        if not row or len(row) < 15:
            continue
        school_code = str(row[2] or "").strip()
        school_id = school_by_code.get(school_code)
        if not school_id:
            continue
        external_id = str(row[10] or "").strip()
        if not external_id:
            continue
        surname = str(row[6] or "").strip()
        name = str(row[7] or "").strip()
        patronymic = str(row[8] or "").strip()
        full_name = " ".join(part for part in (surname, name, patronymic) if part) or external_id
        score = row[14] if row[14] is not None else row[13]
        task_mask = str(row[11] or "").strip()
        parsed_rows.append((school_id, external_id, full_name, float(score or 0), task_mask))
        school_ids.add(school_id)
        ext_ids.add(external_id)

    existing_students = Student.objects.filter(
        school_id__in=school_ids,
        external_id__in=ext_ids,
    ).only("id", "school_id", "external_id")
    student_map = {(s.school_id, s.external_id): s.id for s in existing_students}

    to_create = []
    seen_new_keys = set()
    for school_id, external_id, full_name, _, _ in parsed_rows:
        key = (school_id, external_id)
        if key not in student_map and key not in seen_new_keys:
            seen_new_keys.add(key)
            to_create.append(
                Student(
                    school_id=school_id,
                    external_id=external_id,
                    full_name=full_name,
                    grade="ЕГЭ",
                )
            )
    Student.objects.bulk_create(to_create, batch_size=2000)

    for s in Student.objects.filter(school_id__in=school_ids, external_id__in=ext_ids).only("id", "school_id", "external_id"):
        student_map[(s.school_id, s.external_id)] = s.id

    students_updates = {}
    for school_id, external_id, full_name, _, _ in parsed_rows:
        sid = student_map[(school_id, external_id)]
        students_updates[sid] = full_name
    updates = []
    for sid, full_name in students_updates.items():
        updates.append(Student(id=sid, full_name=full_name, grade="ЕГЭ"))
    Student.objects.bulk_update(updates, ["full_name", "grade"], batch_size=2000)

    final_by_student = {}
    for school_id, external_id, _, total_score, task_mask in parsed_rows:
        key = (school_id, external_id)
        existing = final_by_student.get(key)
        if existing is None or total_score > existing[0]:
            final_by_student[key] = (total_score, task_mask)

    with transaction.atomic():
        ExamResult.objects.filter(exam=exam).delete()
        TaskResult.objects.filter(exam=exam).delete()

        exam_results = []
        task_results = []
        for (school_id, external_id), (total_score, task_mask) in final_by_student.items():
            student_id = student_map[(school_id, external_id)]
            exam_results.append(
                ExamResult(
                    student_id=student_id,
                    exam=exam,
                    total_score=total_score,
                    passed=total_score > 0,
                )
            )
            for idx, value in enumerate(task_mask, start=1):
                task_results.append(
                    TaskResult(
                        student_id=student_id,
                        exam=exam,
                        task_number=idx,
                        value=value,
                    )
                )

        ExamResult.objects.bulk_create(exam_results, batch_size=3000)
        for i in range(0, len(task_results), 10000):
            TaskResult.objects.bulk_create(task_results[i : i + 10000], batch_size=10000)

    print(
        {
            "exam_id": exam.id,
            "subject": exam.subject,
            "students_total": len(student_map),
            "exam_results_created": len(final_by_student),
            "task_results_created": TaskResult.objects.filter(exam=exam).count(),
        }
    )


run()
