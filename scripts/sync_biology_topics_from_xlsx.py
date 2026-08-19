"""
Синхронизация тем ЕГЭ по биологии из Excel в data/ege_2026_enriched.json.
Запуск: python scripts/sync_biology_topics_from_xlsx.py
"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
ENRICHED = ROOT / "data" / "ege_2026_enriched.json"
XLSX = Path.home() / "Desktop" / "Новая папка (5)" / "биология.xlsx"


def norm(value) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"\s+", " ", text).strip()
    text = re.sub(r"(\w)-\s+(\w)", r"\1\2", text)
    return text


def parse_grades(text: str) -> list[int]:
    if not text:
        return []
    grades: set[int] = set()
    for match in re.finditer(r"(\d+)\s*[–-]\s*(\d+)\s*класс", text, flags=re.I):
        lo, hi = sorted((int(match.group(1)), int(match.group(2))))
        for grade in range(lo, hi + 1):
            if 1 <= grade <= 11:
                grades.add(grade)
    for match in re.finditer(r"(\d+)\s*класс", text, flags=re.I):
        grade = int(match.group(1))
        if 1 <= grade <= 11:
            grades.add(grade)
    return sorted(grades)


def read_xlsx_rows() -> dict[int, str]:
    if not XLSX.exists():
        raise SystemExit(f"Файл не найден: {XLSX}")
    wb = load_workbook(XLSX, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    out: dict[int, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        try:
            task_number = int(float(row[0]))
        except (TypeError, ValueError):
            continue
        topic_text = norm(row[1] if len(row) > 1 else "")
        if topic_text:
            out[task_number] = topic_text
    return out


def main() -> None:
    xlsx_map = read_xlsx_rows()
    data = json.loads(ENRICHED.read_text(encoding="utf-8"))
    subjects = data.get("subjects", [])
    biology = next((subject for subject in subjects if subject.get("subject") == "biology"), None)
    if not biology:
        raise SystemExit("В JSON нет subject=biology")

    old_by_task = {int(task["task"]): task for task in biology.get("tasks", []) if task.get("task") is not None}
    merged_tasks: list[dict] = []
    for task_num in sorted(set(old_by_task) | set(xlsx_map)):
        old = old_by_task.get(task_num, {})
        topic_text = xlsx_map.get(task_num, "") or (old.get("topic") or "")
        grades = parse_grades(topic_text) or list(old.get("grade_range") or [])
        item = {
            "task": task_num,
            "topic": topic_text,
            "grade_range": grades,
            "source": "biology_xlsx",
        }
        if old.get("skill"):
            item["skill"] = old["skill"]
        if old.get("topic_oge"):
            item["topic_oge"] = old["topic_oge"]
        if old.get("grade_range_oge"):
            item["grade_range_oge"] = old["grade_range_oge"]
        merged_tasks.append(item)

    biology["tasks"] = merged_tasks
    ENRICHED.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"OK: обновлено {len(merged_tasks)} заданий по биологии в {ENRICHED}")


if __name__ == "__main__":
    main()
